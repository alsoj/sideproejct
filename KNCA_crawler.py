import sys
from PyQt6.QtWidgets import *
from PyQt6 import uic

# 어플리케이션 패키지
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import datetime
from time import sleep

# 전역변수 - 파일 관련
FILE_PATH = './output/'
FILE_PREFIX = 'KNCA_'
FILE_SUFFIX = '.xlsx'
FILE_NAME = ''

# 전역변수 - 재시도 관련
SLEEP_TIME = 10
RETRY_CNT = 5

# 전역변수 - 백그라운드 실행 관련
BACKGROUND_YN = 'Y'

# 전역변수 - 크롤링 대상 주소
HOME_URL = 'http://kcna.kp/kp'
RECENT_URL = 'http://kcna.kp/kp/category/articles/q/1ee9bdb7186944f765208f34ecfb5407.kcmsf'

ROWNUM = 0
START_DATE = ''
END_DATE = ''

form_class = uic.loadUiType("KNCA_crawler.ui")[0]

class KNCA_Window(QMainWindow, form_class):
  def __init__(self):
    super().__init__()
    self.setupUi(self)
    self.btn_start.clicked.connect(self.btn_start_clicked)

  def btn_start_clicked(self):
    global BACKGROUND_YN
    BACKGROUND_YN = "Y" if self.chk_background.isChecked() > 0 else "N"

    self.log_text_browser.clear()
    self.log_text_browser.append("#############################################")
    self.log_text_browser.append("크롤링 작업 시작, 백그라운드 실행 : " + BACKGROUND_YN)
    self.log_text_browser.append("#############################################")
    QApplication.processEvents()

    start_ymd = self.input_start_ymd.date().toPyDate()
    end_ymd = self.input_end_ymd.date().toPyDate()
    str_start_ymd = start_ymd.strftime("%Y%m%d")
    str_end_ymd = end_ymd.strftime("%Y%m%d")
    self.log_text_browser.append("시작일자 :" + str_start_ymd + " / 종료일자 :" + str_end_ymd)
    QApplication.processEvents()

    if check_input_valid(start_ymd, end_ymd):
      global FILE_NAME
      FILE_NAME = FILE_PREFIX + str_start_ymd + '_' + str_end_ymd + FILE_SUFFIX
      self.log_text_browser.append("생성 파일명 : " + FILE_NAME)
      QApplication.processEvents()

      create_excel()  # 엑셀파일 생성

      self.log_text_browser.append(" ")
      self.log_text_browser.append("크롬 브라우저가 실행 중입니다.")
      QApplication.processEvents()

      # 백그라운드 실행 세팅
      options = webdriver.ChromeOptions()
      if BACKGROUND_YN == "Y":
        options.add_argument("headless")

      browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

      try :

        connect(browser, HOME_URL)
        connect(browser, RECENT_URL)

        isGo = True
        while isGo:
          isGo = crawl_list_page(self, browser, start_ymd, end_ymd)
          go_next_page(browser)

      except Exception as e:
        self.log_text_browser.append(" ")
        self.log_text_browser.append("#############################################")
        self.log_text_browser.append("크롤링 작업 오류 발생")
        self.log_text_browser.append(str(e))
        self.log_text_browser.append("#############################################")
        QApplication.processEvents()

      finally:
        self.log_text_browser.append(" ")
        self.log_text_browser.append("#############################################")
        self.log_text_browser.append("크롤링 작업 종료")
        self.log_text_browser.append("#############################################")
        browser.quit()
        QApplication.processEvents()

    else :
      self.log_text_browser.append("오류 발생 : 종료일자는 시작일자보다 크거나 같아야 합니다.")

def check_input_valid(start_ymd, end_ymd):
  if end_ymd < start_ymd:
    return False
  else:
    return True

def create_excel():
  wb = Workbook()
  ws = wb.active

  sub = ['번호','게재일자','기사 제목','(내용1)','내용2','내용3','기사 내용']
  for kwd, j in zip(sub, list(range(1, len(sub) + 1))):
    ws.cell(row=1, column=j).value = kwd

  wb.save(FILE_PATH + FILE_NAME)


def write_excel(append_row):
  wb = load_workbook(FILE_PATH + FILE_NAME, data_only=True)
  ws = wb.active
  ws.append(append_row)
  wb.save(FILE_PATH + FILE_NAME)


def crawl_list_page(self, browser, start_ymd, end_ymd):
  cur_url = browser.current_url
  retries = RETRY_CNT
  while (retries > 0):
    try:
      browser.get(cur_url)
      article_link = browser.find_element(by=By.CLASS_NAME, value='article-link')
      break
    except Exception as e:
      retries = retries - 1
      sleep(SLEEP_TIME)
      print("목록 페이지 로딩 중 응답이 없어 재시도 합니다. 남은 재시도 회수 : ", retries)
      print(e)
      continue

  article_link = browser.find_element(by=By.CLASS_NAME, value='article-link')
  li_tags = article_link.find_elements(by=By.TAG_NAME, value='li')
  for li_tag in li_tags:
    isTarget, isGo, targetUrl, pub_date = check_target(li_tag, start_ymd, end_ymd)
    if isTarget:
      browser.switch_to.new_window('tab')
      connect(browser, targetUrl)
      title, pre_content1, pre_content2, pre_content3, content = get_detail_info(browser)
      global ROWNUM
      ROWNUM += 1

      self.log_text_browser.append(str(ROWNUM) + "번 기사 크롤링 진행 중")
      QApplication.processEvents()

      write_excel([ROWNUM, pub_date, title, pre_content1, pre_content2, pre_content3, content])
      browser.close()
      browser.switch_to.window(browser.window_handles[0])

  return isGo

def check_target(li_tag, start_date, end_date):
  detail_url = li_tag.find_element(by=By.TAG_NAME, value='a').get_attribute('href')
  publish_date = li_tag.find_element(by=By.CLASS_NAME, value='publish-time')

  date = publish_date.text.replace("[", "").replace("]", "").replace("주체", "").split(".")
  year = int(trans_year(date[0]))
  month = int(date[1])
  day = int(date[2])

  pub_date = datetime.date(year, month, day)
  if start_date <= pub_date and pub_date <= end_date:
    isTarget = True
    isGo = True
  elif pub_date < start_date:
    isTarget = False
    isGo = False
  elif pub_date > end_date:
    isTarget = False
    isGo = True
  return isTarget, isGo, detail_url, pub_date


def trans_year(north_year):
  return 1911 + int(north_year)


def get_detail_info(browser):
  cur_url = browser.current_url
  retries = RETRY_CNT
  while (retries > 0):
    try:
      browser.get(cur_url)
      title = browser.find_element(by=By.CLASS_NAME, value='article-main-title').text
      break
    except Exception as e:
      retries = retries - 1
      sleep(SLEEP_TIME)
      print("상세 페이지 로딩 중 응답이 없어 재시도 합니다. 남은 재시도 회수 : ", retries)
      print(e)
      continue

  title = browser.find_element(by=By.CLASS_NAME, value='article-main-title').text
  content_wrapper = browser.find_element(by=By.CLASS_NAME, value='content-wrapper')
  p_tags = content_wrapper.find_elements(by=By.TAG_NAME, value='p')

  pre_content1 = ''
  pre_content2 = ''
  pre_content3 = ''
  content = ''
  for p_tag in p_tags:
    text = p_tag.text
    if text.startswith('(') and text.endswith(')'):
      text = text.replace("(", "").replace(")", "")
      split_text = text.split(' ')
      pre_content1 = split_text[0]
      pre_content2 = split_text[1] + ' ' + split_text[2]
      pre_content3 = split_text[3]
    else:
      content += text + '\r\n'

  if len(content) >= 32000:
    content = content[0:31980] + '\r\n (이하 절단)'

  return title, pre_content1, pre_content2, pre_content3, content


def go_next_page(browser):
  next_btn = browser.find_element(by=By.CLASS_NAME, value='next-ctrl').find_element(by=By.TAG_NAME, value='a')
  retries = RETRY_CNT
  while (retries > 0):
    try:
      browser.execute_script(next_btn.get_attribute('href'))
      article_link = browser.find_element(by=By.CLASS_NAME, value='article-link')
      break
    except Exception as e:
      retries = retries - 1
      sleep(SLEEP_TIME)
      print("다음 페이지 이동 중 응답이 없어 재시도 합니다. 남은 재시도 회수 : ", retries)
      print(e)
      continue


def connect(browser, url):
  retries = RETRY_CNT
  while (retries > 0):
    try:
      browser.get(url)
      print("접속 성공 : ", url)
      break
    except Exception as e:
      retries = retries - 1
      sleep(SLEEP_TIME)
      print("응답이 없어 재시도 합니다. 남은 재시도 회수 : ", retries)
      print(e)
      continue


if __name__ == "__main__":
  app = QApplication(sys.argv)
  KNCA = KNCA_Window()
  KNCA.show()
  app.exec()