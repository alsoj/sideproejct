import sys
from PyQt6.QtWidgets import *
from PyQt6 import uic

# 어플리케이션 패키지
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

import datetime
from time import sleep
import re

import PIL
import io
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 전역변수 - 파일 관련
FILE_PATH = './output/'
FILE_SUFFIX = '.xlsx'
FILE_NAME = ''

ROOT_URL = ''
MEDIA_NAME = ''
# AD_INFO_LIST = []
AD_INFO_DICT = {}
AD_CLASS_DICT = {}
LANDING_CLASS_DICT = {}
# AD_INFO_SET = set()
LANDING_INFO_SET = set()
DEVICE_LIST = []
REPEAT_CNT = 0
TARGET_URL = ''

class Ad:
  def __init__(self, media, device, thumb, text, repeat_cnt, ad_cnt, landing_title, ad_url, landing_url):
    self.media = media
    self.device = device
    self.thumb = thumb
    self.text = text
    self.repeat_cnt = repeat_cnt
    self.ad_cnt = ad_cnt
    self.landing_title = landing_title
    self.ad_url = ad_url
    self.landing_url = landing_url

  def add_cnt(self):
    self.ad_cnt += 1

form_class = uic.loadUiType("ad_crawler.ui")[0]

class AdCrawler_Window(QMainWindow, form_class):
  def __init__(self):
    super().__init__()
    self.setupUi(self)
    self.btn_start.clicked.connect(self.btn_start_clicked)

  # 로그 삭제
  def clear_log(self):
    self.log_browser.clear()

  # 로그 기록
  def set_log_text(self, text):
    self.log_browser.append(text)
    QApplication.processEvents()

  # 디바이스 종류 조회
  def set_device_type(self):
    global DEVICE_LIST
    DEVICE_LIST.clear()
    if self.radio_both.isChecked():
      DEVICE_LIST.append("PC")
      DEVICE_LIST.append("MOBILE")
    elif self.radio_pc.isChecked():
      DEVICE_LIST.append("PC")
    elif self.radio_mobile.isChecked():
      DEVICE_LIST.append("MOBILE")

  # 필수 입력 값 체크
  def check_input_valid(self):
    global REPEAT_CNT
    REPEAT_CNT = self.spin_cnt.value()
    global TARGET_URL
    TARGET_URL = self.edit_url.text()
    if REPEAT_CNT <= 0 or len(TARGET_URL) == 0:
      self.set_log_text("필수 입력 값이 입력되지 않았습니다.")
      self.set_log_text("사이트 URL과 반복횟수를 확인 해주세요.")
      return False
    else:
      return True

  # 작업 시작 버튼 클릭
  def btn_start_clicked(self):
    self.btn_start.setEnabled(False)
    self.clear_log()

    # 필수 입력 값 체크 & 반복회수, 타겟 URL 세팅
    if self.check_input_valid() is not True:
      self.btn_start.setEnabled(True)
      return False

    # 디바이스 타입 세팅
    self.set_device_type()
    self.progress_bar.setValue(0)

    # 전역 변수 세팅 및 초기화
    set_global_variables()

    # 엑셀 파일 생성
    create_excel()

    self.set_log_text(f"매체명 : {MEDIA_NAME}")
    self.set_log_text(f"파일명 : {FILE_NAME}")

    for loop_index, device in enumerate(DEVICE_LIST, start=1):
      self.progress_bar.setMaximum(REPEAT_CNT * len(DEVICE_LIST))
      self.set_log_text("#######################################")
      self.set_log_text("크롤링 작업 시작_" + datetime.datetime.now().strftime("%Y년%m월%d일 %H시%M분"))
      self.set_log_text(f" - 반복회수 : {REPEAT_CNT}")
      self.set_log_text(f" - 기기종류 : {DEVICE_LIST} 중 {device}")
      self.set_log_text("#######################################")

      global AD_CLASS_DICT
      AD_CLASS_DICT.clear()
      global LANDING_CLASS_DICT
      LANDING_CLASS_DICT.clear()

      browser = get_browser(self, device)
      browser.set_page_load_timeout(10)
      browser.get(TARGET_URL)

      for i in range(1, REPEAT_CNT+1):
        try:
          set_crawl_init()  # 크롤링 정보 초기화
          browser.get(TARGET_URL)
          browser_scroll_down(browser)  # 브라우저 스크롤 다운
          crawl_ad(browser)  # 크롤링 진행(웹페이지에 존재하는 a 태그를 추출)

          # for ad_info in AD_INFO_LIST:
          for ad_info in AD_INFO_DICT.values():
            if ad_info['url'] in AD_CLASS_DICT:  # 광고 url를 key로 해서 관리
              # 이미 존재하는 url이라면 cnt +1
              AD_CLASS_DICT[ad_info['url']].add_cnt()
            else:
              # 새로운 url이라면 landing 정보 추가해서 생성
              landing_info = get_landing_info(browser, ad_info)
              if is_not_dup(landing_info['url']):
                if landing_info['url'] in LANDING_CLASS_DICT:
                  LANDING_CLASS_DICT[landing_info['url']].add_cnt()
                else:
                  ad_class = Ad(MEDIA_NAME, device, get_image(ad_info['image']), ad_info['text'], REPEAT_CNT, 1, landing_info['title'], ad_info['url'], landing_info['url'])
                  AD_CLASS_DICT[ad_info['url']] = ad_class
                  LANDING_CLASS_DICT[landing_info['url']] = ad_class

        except Exception as e:
          pass
        finally:
          self.set_log_text(f"{i}회 진행 완료")
          self.progress_bar.setValue(i * loop_index)
          QApplication.processEvents()

      save_excel()
      browser.quit()
      self.btn_start.setEnabled(True)
      self.set_log_text("#######################################")
      self.set_log_text("크롤링 작업 완료_" + datetime.datetime.now().strftime("%Y년%m월%d일 %H시%M분"))
      self.set_log_text("#######################################")

# 크롬 브라우저 로드
def get_browser(self, device):
  self.set_log_text("크롬 브라우저 로딩 중 입니다.")

  options = webdriver.ChromeOptions()
  options.add_argument("headless")

  if device == 'MOBILE':
    mobile_emulation = {"deviceName": "iPhone X"}
    options.add_experimental_option("mobileEmulation", mobile_emulation)

  browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
  browser.maximize_window()

  self.set_log_text("크롬 브라우저 로딩 완료 되었습니다.")
  return browser

def browser_scroll_down(browser):
  scroll_from = 0
  scroll_to = 200
  scroll_height = browser.execute_script("return document.body.scrollHeight")
  sleep(3)
  while (scroll_to < scroll_height):
    browser.execute_script(f"window.scrollTo({scroll_from},{scroll_to})")
    scroll_height = browser.execute_script("return document.body.scrollHeight")
    scroll_from += 200
    scroll_to += 200

# 브라우저 Tab close
def close_new_tabs(browser):
  tabs = browser.window_handles
  while len(tabs) != 1:
    browser.switch_to.window(tabs[1])
    browser.close()
    tabs = browser.window_handles
  browser.switch_to.window(tabs[0])

# 전역변수 세팅 및 초기화
def set_global_variables():
  global ROOT_URL
  ROOT_URL = get_root_url()
  global MEDIA_NAME
  MEDIA_NAME = get_media_name()

  device = ''
  if 'PC' in DEVICE_LIST:
    device += '_PC'
  if 'MOBILE' in DEVICE_LIST:
    device += '_MOBILE'
  global FILE_NAME
  FILE_NAME = datetime.datetime.now().strftime("%Y%m%d%H%M%S") + f"_{MEDIA_NAME}{device}" + FILE_SUFFIX

def set_crawl_init():
  # global AD_INFO_LIST
  # AD_INFO_LIST = []
  global AD_INFO_DICT
  AD_INFO_DICT.clear()
  # global AD_INFO_SET
  # AD_INFO_SET.clear()
  global LANDING_INFO_SET
  LANDING_INFO_SET.clear()

# 엑셀 파일 생성
def create_excel():
  wb = Workbook()
  ws = wb.active

  sub = ['매체','기기','썸네일','텍스트','새로고침','노출','타이틀','광고URL','랜딩페이지URL']
  for kwd, j in zip(sub, list(range(1, len(sub) + 1))):
    ws.cell(row=1, column=j).value = kwd

  ws.column_dimensions['C'].width = 25
  ws.column_dimensions['D'].width = 25
  ws.column_dimensions['G'].width = 40
  ws.column_dimensions['H'].width = 50
  ws.column_dimensions['I'].width = 50

  wb.save(FILE_PATH + FILE_NAME)

# 엑셀 파일 광고 존재 여부 확인 후 업데이트
def check_update_excel(ad_url):
  wb = load_workbook(FILE_PATH + FILE_NAME, data_only=True)
  ws = wb.active

  for index, row in enumerate(ws.rows, start=1):
    if ad_url == row[7].value:
      ws.cell(row=index, column=6).value = row[5].value + 1
      wb.save(FILE_PATH + FILE_NAME)
      return True
  return False

# 엑셀 파일 저장(마지막에 한 번에)
def save_excel():
  wb = load_workbook(FILE_PATH + FILE_NAME, data_only=True)
  ws = wb.active
  rownum = ws.max_row + 1

  for url, ad_class in AD_CLASS_DICT.items():
    ws.row_dimensions[rownum].height = 70
    ws['A'+str(rownum)] = ad_class.media
    ws['B'+str(rownum)] = ad_class.device
    image = ad_class.thumb
    if type(image) != str:
      image.height = 85
      image.width = 200
      ws.add_image(image, 'C'+str(rownum))
    else:
      ws['C'+str(rownum)] = image
    ws['D'+str(rownum)] = ad_class.text
    ws['E'+str(rownum)] = ad_class.repeat_cnt
    ws['F'+str(rownum)] = ad_class.ad_cnt
    ws['G'+str(rownum)] = ad_class.landing_title
    ws['H'+str(rownum)] = ad_class.ad_url
    ws['I'+str(rownum)] = ad_class.landing_url
    rownum += 1

  wb.save(FILE_PATH + FILE_NAME)

# 엑셀 파일 insert (존재 하지 않는 경우)
def insert_excel(excel_input):
  wb = load_workbook(FILE_PATH + FILE_NAME, data_only=True)
  ws = wb.active

  rownum = ws.max_row+1
  ws.row_dimensions[rownum].height = 70
  ws['A'+str(rownum)] = excel_input['media']
  ws['B'+str(rownum)] = excel_input['device']
  image = excel_input['thumb']
  if type(image) != str:
    image.height = 85
    ws.add_image(image, 'C'+str(rownum))
  else:
    ws['C'+str(rownum)] = image
  ws['D'+str(rownum)] = excel_input['text']
  ws['E'+str(rownum)] = excel_input['repeat_cnt']
  ws['F'+str(rownum)] = excel_input['ad_cnt']
  ws['G'+str(rownum)] = excel_input['langding_title']
  ws['H'+str(rownum)] = excel_input['ad_url']
  ws['I'+str(rownum)] = excel_input['landing_url']

  wb.save(FILE_PATH + FILE_NAME)

# 메인 URL 추출
def get_root_url():
  return TARGET_URL.split('/')[2]

# 매체 명 추출
def get_media_name():
  root_url = TARGET_URL.split('/')[2]
  media_name = root_url.replace("www.", "")
  media_name = media_name.replace(".com", "")
  media_name = media_name.replace(".co.kr", "")
  media_name = media_name.replace(".kr", "")
  media_name = media_name.replace(".net", "")
  media_name = media_name.replace("m.", "")
  media_name = media_name.replace("mobile.", "")

  return media_name

# 광고 여부 판단
def isAd(href):
  if href is None:
    return False
  elif 'ad' in href:
    if 'whythisad' in href \
      or 'adsense/support' in href \
      or 'reader' in href \
      or 'header' in href\
      or 'info' in href:
      return False
    else:
      return True
  elif ROOT_URL in href:
    return False
  elif 'javascript' in href \
    or MEDIA_NAME in href \
    or 'youtube' in href \
    or 'weather' in href \
    or 'facebook' in href \
    or 'twitter' in href \
    or 'newsstand.naver' in href \
    or 'media.naver' in href:
    return False
  else:
    return True

# 광고 정보 추출(from a 태그)
def get_ad_info(a_tag):
  ad_info = {}
  text = ''
  image = ''
  url = ''

  try:
    text = a_tag.text
    url = a_tag.get_attribute('href')
    images = a_tag.find_elements(by=By.TAG_NAME, value='img')
    onclick_attr = a_tag.get_attribute('onclick')

    # 이미지의 크기가 1kb 초과인 것들은 썸네일로 간주
    if len(images) > 0:
      image_src = images[0].get_attribute('src') if images[0].get_attribute('src') is not None else ''
      if len(image_src) > 0:
        image_content = requests.get(image_src).content
        if len(image_content) > 1024:
          image = image_src

    # onclick에 실제 연결 url이 있는 경우
    if onclick_attr is not None and "(" in onclick_attr and ")" in onclick_attr and "/" in onclick_attr:
      onclick = re.findall('\(([^)]+)', onclick_attr)[0].replace("\"", "").replace("'", "")
      onclick = onclick.split("//")[-1]
      onclick = 'https://' + onclick
      if len(onclick) > 0:
        url = onclick
  except Exception as e:
    print("get_ad_info 오류 : " + str(e))
    pass

  finally:
    ad_info['text'] = text
    ad_info['image'] = image
    ad_info['url'] = url
    return ad_info

# 랜딩 페이지 정보 추출(from ad_info)
def get_landing_info(browser, ad_info):
  landing_info = {}
  landing_title = ''
  landing_url = ''

  try:
    # res = requests.get(ad_info['url'])
    # soup = BeautifulSoup(res.content, "html.parser")
    # landing_title = soup.find("title").get_text()
    # landing_url = res.url
    browser.switch_to.new_window('tab')
    browser.get(ad_info['url'])
    landing_title = browser.title
    landing_url = browser.current_url

  except Exception as e:
    print("get_landing_info 오류 " + e)
    pass

  finally:
    landing_info['title'] = landing_title
    landing_info['url'] = landing_url
    close_new_tabs(browser)

  return landing_info

# 이미지 추출
def get_image(image_url):
  img = ''
  try:
    if len(image_url) > 0:
      headers = {"user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"}
      http = urllib3.PoolManager()
      r = http.request('GET', image_url, headers=headers)
      image_file = io.BytesIO(r.data)
      img = Image(image_file)
  except Exception as e:
    print("get_image 오류 : " + e)
    pass

  finally:
    return img

# 페이지에 존재하는 모든 광고 수집
def crawl_ad(browser):
  try:
    # 해당 iframe에서 광고 추출
    a_tags = browser.find_elements(by=By.TAG_NAME, value='a')

    for a_tag in a_tags:
      if isAd(a_tag.get_attribute('href')):
        ad_info = get_ad_info(a_tag)
        set_ad_info(ad_info)

    # 하위 iframe 조회 및 step into
    iframes = browser.find_elements(by=By.TAG_NAME, value='iframe')

    # 브라우저에 iframe이 있는 경우 순서대로 iframe으로 in
    if iframes is not None and len(iframes) > 0:
      for i in range(len(iframes)):
        browser.switch_to.frame(i)  # iframe으로 IN
        crawl_ad(browser)
        browser.switch_to.parent_frame()  # iframe에서 OUT

  except Exception as e:
    print("crawl_ad 오류 : ", e)
    pass

# ad info 세팅
def set_ad_info(ad_info):
  # global AD_INFO_SET
  global AD_INFO_DICT

  try:
    if ad_info is None:
      return

    # 이미 수집된 url 중 이미지/텍스트 없으면 보충
    if ad_info['url'] in AD_INFO_DICT:
      if len(AD_INFO_DICT[ad_info['url']]['image']) == 0 and len(ad_info['image']) > 0:
        AD_INFO_DICT[ad_info['url']]['image'] = ad_info['image']
      if len(AD_INFO_DICT[ad_info['url']]['text']) == 0 and len(ad_info['text']) > 0:
        AD_INFO_DICT[ad_info['url']]['text'] = ad_info['text']

    elif (len(ad_info['text']) > 0 or len(ad_info['image']) > 0 or 'ad' in ad_info['url']) \
      and ad_info['url'] not in AD_INFO_DICT:
      AD_INFO_DICT[ad_info['url']] = ad_info

      # global AD_INFO_LIST
      # AD_INFO_LIST.append(ad_info)
      # AD_INFO_SET.add(ad_info['url'])

  except Exception as e:
    print("set_ad_info 오류 : " + str(e))

# 엑셀 입력 값 세팅
def get_excel_input(ad_info, landing_info, repeat_cnt):
  excel_input = {}
  excel_input['media'] = MEDIA_NAME
  excel_input['device'] = 'PC'
  excel_input['thumb'] = get_image(ad_info['image'])
  excel_input['text'] = ad_info['text']
  excel_input['repeat_cnt'] = repeat_cnt
  excel_input['ad_cnt'] = 1
  excel_input['langding_title'] = landing_info['title']
  excel_input['ad_url'] = ad_info['url']
  excel_input['landing_url'] = landing_info['url']
  return excel_input

# 랜딩 페이지 기 등록 여부 확인
def is_not_dup(landing_url):
  global LANDING_INFO_SET
  if landing_url not in LANDING_INFO_SET:
    LANDING_INFO_SET.add(landing_url)
    return True
  else:
    return False


if __name__ == "__main__":
  app = QApplication(sys.argv)
  AdCrawler = AdCrawler_Window()
  AdCrawler.show()
  app.exec()