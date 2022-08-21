import requests
import json
import datetime

from openpyxl import Workbook, load_workbook

NOW = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
FILE_PATH = './output/'
FILE_NAME = f'{NOW}_viewer추출.xlsx'

# 엑셀 파일 생성
def create_excel():
    wb = Workbook()
    ws = wb.active

    col_names = ['no', 'userName', 'fullName', 'visitsTotal']
    for kwd, j in zip(col_names, list(range(1, len(col_names) + 1))):
        ws.cell(row=1, column=j).value = kwd

    wb.save(FILE_PATH + FILE_NAME)

# 엑셀 입력
def write_excel(append_row):
    wb = load_workbook(FILE_PATH + FILE_NAME, data_only=True)
    ws = wb.active

    ws.append(append_row)
    wb.save(FILE_PATH + FILE_NAME)


if __name__ == '__main__':
    test_json = '{"results":[{"userName":"jp29028","fullName":"Park Sejin[ 朴世振 ]","avatarPath":"/images/icons/profilepics/default.svg","formattedLastVisitTime":"8월 21, 2022","visitsTotal":46,"knownUser":true},{"userName":"lw12787","fullName":"Park Sejin[ 朴世振 ]","avatarPath":"/images/icons/profilepics/default.svg","formattedLastVisitTime":"8월 19, 2022","visitsTotal":1,"knownUser":true},{"userName":"jp30795","fullName":"Park Sejin[ 朴世振 ]","avatarPath":"/images/icons/profilepics/default.svg","formattedLastVisitTime":"8월 19, 2022","visitsTotal":2,"knownUser":true},{"userName":"jp29964","fullName":"Park Sejin[ 朴世振 ]","avatarPath":"/images/icons/profilepics/default.svg","formattedLastVisitTime":"8월 19, 2022","visitsTotal":29,"knownUser":true},{"userName":"jp30689","fullName":"Park Sejin[ 朴世振 ]","avatarPath":"/images/icons/profilepics/default.svg","formattedLastVisitTime":"8월 19, 2022","visitsTotal":11,"knownUser":true},{"userName":"jpz4378","fullName":"Park Sejin[ 朴世振 ]","avatarPath":"/images/icons/profilepics/default.svg","formattedLastVisitTime":"8월 18, 2022","visitsTotal":7,"knownUser":true}]}'

    try:
        rownum = 0
        input_site_url = input("추출할 URL을 입력해주세요 : ")
        input_site_url = input_site_url.strip()

        res = requests.get(input_site_url)
        data = json.loads(res.text)
        results = data['results']
        for result in results:
            rownum += 1
            write_excel(rownum, result['userName'], result['fullName'], result['visitsTotal'])

    except Exception as e:
        print("오류 발생 " + e)

