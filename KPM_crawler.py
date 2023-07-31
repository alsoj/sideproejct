import sys
import os
import platform

from PyQt6.QtCore import QDate, QThread
from PyQt6.QtWidgets import *

from PyQt6.QtWidgets import QMainWindow
from datetime import datetime, timedelta
from PyQt6 import uic

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller

from openpyxl import Workbook, load_workbook

from time import sleep
from selenium.common.exceptions import NoSuchElementException

import unicodedata
import re


import logging
logger = logging.getLogger()  # 로그 생성
logger.setLevel(logging.DEBUG)  # 로그의 출력 기준 설정
formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')  # log 출력 형식

# log를 파일에 출력
file_handler = logging.FileHandler('KPM_crawler.log')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# 전역변수 - 파일 관련
FILE_PATH = './output/'
FILE_PREFIX = 'KPM_'
FILE_SUFFIX = '.xlsx'
FILE_NAME = ''

# 전역변수 - 재시도 관련
SLEEP_TIME = 60
RETRY_CNT = 10

# ID/PW 세팅
f = open("KPM_account.txt", 'r')
lines = f.readlines()
ID, PW = lines[0].strip(), lines[1].strip()
f.close()

form_class = uic.loadUiType("KPM_crawler.ui")[0]
class KPMCrawler(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.news_worker = NewsWorker(self)

        # 날짜 기본 세팅
        self.input_start_ymd.setDate(QDate(datetime.now()-timedelta(days=7)))
        self.input_end_ymd.setDate(QDate(datetime.now()))
        self.scroll_bar = self.log_browser.verticalScrollBar()

        # 작업 시작 클릭 시
        self.btn_start.clicked.connect(self.crawl_news)

    def debug(self, text):  # 디버그 로그 출력
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f"[{now}] " + text)
        self.scroll_bar.setValue(self.scroll_bar.maximum())
        logger.debug(text)

    def info(self, text):  # 정보 로그 출력
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f'<p style="color:green">[{now}] {text}</p>')
        self.scroll_bar.setValue(self.scroll_bar.maximum())
        logger.info(text)

    def error(self, text):  # 에러 로그 출력
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f'<p style="color:red">[{now}] {text}</p>')
        self.scroll_bar.setValue(self.scroll_bar.maximum())
        logger.error(text)

    def crawl_news(self):
        logger.debug("작업이 시작되었습니다.")
        try:
            if self.input_end_ymd.date().toPyDate() < self.input_start_ymd.date().toPyDate():
                self.error("오류 발생 : 종료일자는 시작일자보다 크거나 같아야 합니다.")
                logger.error("오류 발생 : 종료일자는 시작일자보다 크거나 같아야 합니다.")
            else:
                background_yn = "Y" if self.chk_background.isChecked() > 0 else "N"
                options = get_option(background_yn)

                # chromedriver_autoinstaller.install(True)
                # self.browser = webdriver.Chrome(options=options)

                self.browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
                # self.browser = webdriver.Chrome(executable_path='./chromedriver', options=options)
                logger.debug("크롬 브라우저가 기동되었습니다.")

                # if 'macOS' in platform.platform():
                #     self.browser = webdriver.Chrome(executable_path='/Users/alsoj/Workspace/kmong/ipynb/chromedriver113', options=options)
                # else:
                #     self.browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

        except Exception as e:
            logger.error("오류 발생 : " + str(e))
            self.browser = webdriver.Chrome(executable_path='./chromedriver', options=options)
            logger.debug("크롬 브라우저가 기동되었습니다.")
        finally:
            self.news_worker.start()

    def closeEvent(self, QCloseEvent):
        if hasattr(self, 'browser'):
            self.browser.quit()

class NewsWorker(QThread):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        try:
            self.parent.btn_start.setEnabled(False)
            self.parent.info("조선언론정보기지 크롤링 START")

            start_ymd = self.parent.input_start_ymd.date().toPyDate()
            end_ymd = self.parent.input_end_ymd.date().toPyDate()
            str_start_ymd = start_ymd.strftime("%Y%m%d")
            str_end_ymd = end_ymd.strftime("%Y%m%d")

            # 엑셀 세팅
            global FILE_NAME
            FILE_NAME = FILE_PREFIX + str_start_ymd + "_" + str_end_ymd + FILE_SUFFIX

            create_excel()
            wb = load_workbook(FILE_PATH + FILE_NAME, data_only=True)
            ws = wb.active

            ###################################
            # 메인 로직 시작
            ###################################
            login_url = 'https://dprkmedia.com/login/'

            target_news = ""
            if self.parent.check_RD.isChecked():
                target_news += "&media[]=223"
            if self.parent.check_MH.isChecked():
                target_news += "&media[]=225"
            if self.parent.check_MJ.isChecked():
                target_news += "&media[]=227"
            if self.parent.check_PT.isChecked():
                target_news += "&media[]=51338"
            if self.parent.check_JS.isChecked():
                target_news += "&media[]=51174"
            search_call_url = f'https://dprkmedia.com/search-ko/?target=post&period=select&period_any[]={start_ymd}&period_any[]={end_ymd}{target_news}&limit=1000'

            self.connect_url(login_url)

            if login(self.parent.browser):
                self.parent.debug("로그인에 성공했습니다.")
            elif login(self.parent.browser):
                self.parent.debug("로그인에 성공했습니다.")
            elif login(self.parent.browser):
                self.parent.debug("로그인에 성공했습니다.")
            else:
                self.parent.error("로그인에 실패했습니다. 잠시 후에 다시 시도해주세요.")

            self.connect_url(search_call_url)
            article_list = get_article_list(self.parent.browser)
            article_cnt = len(article_list)
            rownum = 0
            self.parent.info(f"크롤링 대상 기사 개수 : {article_cnt}")
            for article in article_list:
                detail_url = article.find_element(By.TAG_NAME, value='a').get_attribute('href')
                self.parent.browser.switch_to.new_window('tab')
                self.connect_url(detail_url)
                detail_info = get_detail_info(self.parent.browser)

                ws.append(detail_info)
                wb.save(FILE_PATH + FILE_NAME)
                close_new_tabs(self.parent.browser)
                rownum += 1
                if rownum % 10 == 0:
                    self.parent.debug(f"기사 크롤링 진행 중 [{rownum} / {article_cnt}개 완료]")

            self.parent.info("조선언론정보기지 크롤링 END")
        except Exception as e:
            self.parent.error(f"크롤링 진행 중 오류 발생 : {str(e)}")
        finally:
            self.parent.browser.quit()
            self.parent.btn_start.setEnabled(True)

    def connect_url(self, url):
        retries = RETRY_CNT
        connected = False
        while retries > 0:
            try:
                self.parent.browser.get(url)
                connected = True
                break
            except Exception as e:
                retries -= 1
                self.parent.debug("응답이 없어 재시도 합니다. 남은 재시도 회수 : " + str(retries))
                sleep(SLEEP_TIME)
                continue
        return connected

# 전역 함수
def get_option(background_yn):
    # 백그라운드 실행 세팅
    options = webdriver.ChromeOptions()
    if background_yn == "Y":
        options.add_argument("headless")
    return options

def create_excel():
    # 엑셀 생성
    wb = Workbook()
    ws = wb.active

    # 제목 적기
    sub = ['기사제목', '신문명', '기사일자', '기사내용', '기자명', '기사분류']
    for kwd, j in zip(sub, list(range(1, len(sub) + 1))):
        ws.cell(row=1, column=j).value = kwd

    if not os.path.isdir(FILE_PATH):
        os.mkdir(FILE_PATH)

    wb.save(FILE_PATH+FILE_NAME)

# 로그인
def login(browser):
    input_id = browser.find_element(by=By.ID, value='user_login-684434')
    input_pw = browser.find_element(by=By.ID, value='user_password-684434')
    button_login = browser.find_element(by=By.ID, value='um-submit-btn')

    input_id.clear()
    input_id.send_keys(ID)
    input_pw.clear()
    input_pw.send_keys(PW)
    button_login.click()

    # 로그인 여부 확인
    if len(browser.find_elements(by=By.ID, value='user_login-684434')) == 0:
        return True
    else:
        return False

# 기사 리스트 조회
def get_article_list(browser):
    search_result = browser.find_element(by=By.CLASS_NAME, value='search_result')
    article_list = search_result.find_elements(by=By.TAG_NAME, value='article')
    return article_list

# 신규 오픈 탭 닫기
def close_new_tabs(browser):
    tabs = browser.window_handles
    while len(tabs) != 1:
        browser.switch_to.window(tabs[1])
        browser.close()
        tabs = browser.window_handles
    browser.switch_to.window(tabs[0])

# 페이지에서 원하는 값 추출
def find_element(browser, target):
    try:
        if target == 'media_name':
            return browser.find_element(by=By.CLASS_NAME, value='cat_name').text.strip()
        elif target == 'section':
            return browser.find_element(by=By.ID, value='tag_list').text.strip()
        elif target == 'title':
            return browser.find_element(by=By.CLASS_NAME, value='article-title').text.strip().split('\u3000')[0].replace("[JP]","")
        elif target == 'subtitle':
            return '\n' + browser.find_element(by=By.CLASS_NAME, value='article_subtitle').text.strip().replace("[JP]","")
        elif target == 'content':
            content = browser.find_element(by=By.CLASS_NAME, value='article-content')
            p_lines = content.find_elements(by=By.TAG_NAME, value='p')
            content_text = ''
            for p in p_lines:
                content_text += p.text.strip() + '\n'
            return content_text
        elif target == 'author':
            return browser.find_element(by=By.CLASS_NAME, value='writers').text.strip()
        elif target == 'date':
            date = browser.find_element(by=By.ID, value='post_date').text.strip()
            date = re.search(r'\d{4}년\d*월\d*일', date).group().replace("년","-").replace("월","-").replace("일","").split("-")
            year = date[0]
            month = date[1].zfill(2)
            day = date[2].zfill(2)
            return year + "-" + month + "-" + day
        else:
            return ''
    except NoSuchElementException:
        return ''

def get_detail_info(browser):
    media_name = find_element(browser, 'media_name')
    media_name = unicodedata.normalize('NFKC', media_name)
    section = find_element(browser, 'section')
    section = unicodedata.normalize('NFKC', section)
    main_title = find_element(browser, 'title')
    main_title = unicodedata.normalize('NFKC', main_title)
    subtitle = find_element(browser, 'subtitle')
    subtitle = unicodedata.normalize('NFKC', subtitle)
    title = main_title + subtitle
    title = title.strip()
    content = find_element(browser, 'content')
    content = unicodedata.normalize('NFKC', content)
    author = find_element(browser, 'author')
    author = unicodedata.normalize('NFKC', author)
    date = find_element(browser, 'date')
    date = unicodedata.normalize('NFKC', date)

    return [title, media_name, date, content, author, section]


if __name__ == "__main__":
    logger.debug("프로그램이 실행되었습니다.")
    app = QApplication(sys.argv)
    news_crawler = KPMCrawler()
    news_crawler.show()
    app.exec()