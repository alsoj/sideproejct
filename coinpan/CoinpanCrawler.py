import os
import sys

from PyQt6.QtWidgets import QMainWindow
from datetime import datetime
from PyQt6 import uic

# UI영역
import Config
from PostWorker import PostWorker
from PriceWorker import PriceWorker
from openpyxl import Workbook, load_workbook

form_class = uic.loadUiType(Config.UI_DIR)[0]
class CoinpanCrawler(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.price_worker = PriceWorker(self)
        self.post_worker = PostWorker(self)

        self.btn_price.clicked.connect(self.crawl_price)
        self.btn_browser.clicked.connect(self.execute_browser)
        self.btn_start.clicked.connect(self.click_start)
        self.btn_stop.clicked.connect(self.click_stop)

        self.filename, self.wb, self.ws = None, None, None

        self.debug("코인판 크롤링 프로그램이 실행되었습니다.")

    def debug(self, text):  # 디버그 로그 출력
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f"[{now}] " + text)

    def info(self, text):  # 정보 로그 출력
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f'<p style="color:green">[{now}] {text}</p>')

    def error(self, text):  # 에러 로그 출력
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f'<p style="color:red">[{now}] {text}</p>')

    def crawl_price(self):
        # 실시간 시세 크롤링
        self.price_worker.start()

    def execute_browser(self):
        # self.post_worker.execute_browser()
        self.post_worker.set_function("execute_browser")
        self.post_worker.start()  # 크롬 브라우저 로딩

    def check_page(self):
        try:
            if int(self.edit_page.text().strip()) > 0:
                return True
            else:
                raise Exception()
        except Exception as e:
            self.error("시작 페이지 값을 입력해주세요.")
            return False

    def crawl_post(self):
        if self.radio_recent.isChecked():
            self.post_worker.set_function("crawl_recent")
            self.post_worker.start()  # 최신 크롤링
        else:
            self.post_worker.set_function("crawl_past")
            self.post_worker.start()  # 과거 크롤링

    def click_start(self):
        if not self.check_page():
            return False

        if self.post_worker.browser is None:
            self.error("크롬 브라우저를 실행하여 로그인을 먼저 진행해주세요.")
        else:
            self.post_worker.resume()
            self.btn_start.setEnabled(False)
            self.btn_stop.setEnabled(True)

            # 엑셀 파일 생성 & 로드
            self.filename = create_excel()
            self.wb = load_workbook(self.filename, data_only=True)
            self.ws = self.wb.active
            self.crawl_post()

    def click_stop(self):
        if self.post_worker.browser is None:
            self.error("크롬 브라우저를 실행하여 로그인을 먼저 진행해주세요.")
        else:
            self.post_worker.stop()
            self.btn_start.setEnabled(True)
            self.btn_stop.setEnabled(False)

            # 엑셀 파일 저장
            self.wb.save(self.filename)


# 엑셀 파일 생성
def create_excel():
    wb = Workbook()
    ws = wb.active

    title = ['게시물 번호', '제목', '닉네임', '레벨', '추천', '비추천', '댓글', '조회수', '등록일', '본문', '이미지 주소', '댓글']

    for kwd, j in zip(title, list(range(1, len(title) + 1))):
        ws.cell(row=1, column=j).value = kwd

    filename = Config.OUTPUT_DIR + datetime.now().strftime('%Y%m%d-%H%M%S') + Config.BOARD_FILENAME
    wb.save(filename)
    return filename