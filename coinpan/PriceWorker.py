import os
from PyQt6.QtCore import QThread

from datetime import datetime
from pandas import DataFrame
from time import time
import requests
from openpyxl import Workbook, load_workbook

import Config

class PriceWorker(QThread):
    """
    실시간 시세 크롤링 쓰레드
    """
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        try:
            self.parent.btn_price.setEnabled(False)
            self.parent.info("실시간 시세 추출 시작")
            result = get_price()

            if not os.path.isfile(Config.OUTPUT_DIR+Config.PRICE_FILENAME_FOR_APPEND):
                create_excel()

            append_excel(result)
            # filename = export_excel(result)
            # self.parent.debug("파일 저장 경로 : " + filename)
            self.parent.info("실시간 시세 추출 종료")
            self.parent.btn_price.setEnabled(True)
        except Exception as e:
            print("실시간 시세 크롤링 중 오류 발생", e)


##############################
# Static Method 영역
##############################
def get_price():
    """
    실시간 시세 API 호출
    :return: 실시간 시세 API
    """
    now = round(time())
    response = requests.get(Config.API_URL + str(now))
    data = response.json()
    call_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    prices = DataFrame(
        columns=['순번1', '순번2', '추출 시간', '가상화폐', '거래소', '실시간시세(KRW)', '실시간시세(USD)', '24시간변동액', '24시간변동률', '한국프리미엄',
                 '한국프리미엄(%)', '거래량'])

    for i, exchange in enumerate(Config.API_EXCHANGE):
        if exchange in data['prices']:
            for j, coin in enumerate(Config.API_COIN):
                if coin in data['prices'][exchange]:
                    price = data['prices'][exchange][coin]
                    prices.loc[len(prices)] = [j, i, call_time, coin, exchange,
                                                round(float(price['now_price'] or 0), 2),
                                                round(float(price['now_price_usd'] or 0), 4),
                                                round(float(price['diff_24hr'] or 0), 2),
                                                round(float(price['diff_24hr_percent'] or 0), 2),
                                                round(float(price['korea_premium'] or 0), 2),
                                                round(float(price['korea_premium_percent'] or 0), 2),
                                                round(float(price['units_traded'] or 0), 2)]
    result = prices.sort_values(by=['순번1','순번2'])
    result = result.drop(['순번1','순번2'], axis=1)
    result = result.reset_index(drop=True)

    return result.values.tolist()

def append_excel(result):
    """
    엑셀 append 함수
    :param result: 실시간 시세 list of list
    :return: filename : 경로 및 파일명
    """
    wb = load_workbook(Config.OUTPUT_DIR + Config.PRICE_FILENAME_FOR_APPEND, data_only=True)
    ws = wb.active
    for row in result:
        ws.append(row)
    wb.save(Config.OUTPUT_DIR + Config.PRICE_FILENAME_FOR_APPEND)

# 엑셀 파일 생성
def create_excel():
    wb = Workbook()
    ws = wb.active

    title = ['추출시간', '가상화폐', '거래소', '실시간시세(KRW)', '실시간시세(USD)', '24시간변동액', '24시간변동률', '한국프리미엄',
                 '한국프리미엄(%)', '거래량']

    for kwd, j in zip(title, list(range(1, len(title) + 1))):
        ws.cell(row=1, column=j).value = kwd

    wb.save(Config.OUTPUT_DIR + Config.PRICE_FILENAME_FOR_APPEND)

def export_excel(result):
    """
    엑셀 export 함수
    :param result: 실시간 시세 Dataframe
    :return filename : 경로 및 파일명
    """
    filename = Config.OUTPUT_DIR + datetime.now().strftime('%Y%m%d-%H%M%S') + Config.PRICE_FILENAME
    result.to_excel(filename, index=False)
    return filename