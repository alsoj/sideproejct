from time import sleep

from PyQt6.QtCore import QThread
import json

from selenium.webdriver.common.by import By
import Config
import unicodedata
from openpyxl import load_workbook
from Common import get_app_id, get_short_code
import requests

from instagram import Common


class LikerWorker(QThread):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        # 로그인 된 이후에 실행되어야 함
        code_list = []
        if self.parent.crawl_type == 'recent':
            code_list = get_recent_post(self.parent.browser, self.parent.target_id, self.parent.target_count)
        else:
            code_list = self.get_target_post()

        total_user = {}
        crawl_count = 0
        for index, code in enumerate(code_list):
            if len(code) > 0:
                crawl_count += 1
                sheet_num = index + 1
                Common.info(self.parent.log_browser, f"{sheet_num}번 째 게시글 추출을 시작합니다.")

                # self.parent.info(f"{sheet_num}번 째 게시글 추출을 시작합니다.")
                user_list = self.get_likers(self.parent.browser, code, sheet_num)
                for user in user_list:
                    if user in total_user.keys():
                        total_user[user] += 1
                    else:
                        total_user[user] = 1

        # file_path_name = Config.FILE_PATH + self.parent.file_name
        # common_user = [k for k, v in total_user.items() if v == crawl_count]
        # wb = load_workbook(file_path_name, data_only=True)
        # ws = wb['공통 사용자']
        # rownum = 0
        # for user in common_user:
        #     rownum += 1
        #     ws.append([rownum, user])
        # wb.save(file_path_name)

        if self.parent.crawl_type == 'recent':
            Common.info(self.parent.log_browser, "최근 게시물 크롤링이 종료 되었습니다.")
        else:
            Common.info(self.parent.log_browser, "특정 게시물 크롤링이 종료 되었습니다.")

        self.parent.button_activate(True)

    def get_target_post(self):
        url1 = get_short_code(self.parent.edit_url1.text())
        url2 = get_short_code(self.parent.edit_url2.text())
        url3 = get_short_code(self.parent.edit_url3.text())
        url4 = get_short_code(self.parent.edit_url4.text())
        url5 = get_short_code(self.parent.edit_url5.text())
        url6 = get_short_code(self.parent.edit_url6.text())
        code_list = [url1, url2, url3, url4, url5, url6]
        return code_list

    # 좋아요 추출
    def get_likers(self, browser, short_code, sheet_num):
        end_cursor = ''
        variables = {'shortcode': short_code, 'first': 50, 'after': end_cursor}
        has_next_page = True
        rownum = 0
        file_path_name = Config.FILE_PATH + self.parent.file_name
        wb = load_workbook(file_path_name, data_only=True)
        ws = wb['게시글'+str(sheet_num)]
        user_list = []

        while has_next_page:
            variables['after'] = end_cursor
            json_variables = str(json.dumps(variables))
            url = f'{Config.LIKER_URL}&variables={json_variables}'

            browser.get(url)
            sleep(3)
            content = browser.find_element(by=By.TAG_NAME, value='pre').text
            data = json.loads(content)
            # print(data)

            if 'data' in data:
                has_next_page = data['data']['shortcode_media']['edge_liked_by']['page_info']['has_next_page']
                end_cursor = data['data']['shortcode_media']['edge_liked_by']['page_info']['end_cursor']
                likers = data['data']['shortcode_media']['edge_liked_by']['edges']

                for liker in likers:
                    rownum += 1
                    user_list.append(liker['node']['username'])
                    ws.append([short_code, rownum, liker['node']['username'], unicodedata.normalize('NFC', liker['node']['full_name'])])
                    if rownum % 50 == 0:
                        Common.debug(self.parent.log_browser, f'{rownum}번째 데이터 추출 완료')
                        wb.save(file_path_name)
            else:
                has_next_page = False
                Common.error(self.parent.log_browser, f'{rownum}번째 데이터 추출 완료')

        wb.save(file_path_name)

        return user_list


def get_recent_post(browser, user_id, target_count):
    timeline_url = f'{Config.BASE_URL}' + user_id

    browser.get(timeline_url)
    page_source = browser.page_source
    app_id = get_app_id(page_source)
    csrftoken = browser.get_cookie("csrftoken")['value']

    cookies = browser.get_cookies()
    header_cookie = ''
    for cookie in cookies:
        header_cookie += f"{cookie['name']}={cookie['value']}; "

    # 헤더 세팅
    headers = {'referer': Config.BASE_URL, 'x-csrftoken': csrftoken, 'cookie': header_cookie, 'x-ig-app-id': app_id}
    api_url = f'https://www.instagram.com/api/v1/feed/user/{user_id}/username/?count=12'
    res = requests.get(api_url, headers=headers)
    data = res.json()

    code_list = []
    for i in range(target_count):
        code_list.append(data['items'][i]['code'])

    return code_list
