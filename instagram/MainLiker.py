import os
import sys
from time import sleep

from PyQt6.QtWidgets import *
from PyQt6 import uic

from datetime import datetime
from Common import execute_browser
from openpyxl import Workbook

import Config
from Liker import LikerWorker
from selenium.webdriver.common.by import By

form_class = uic.loadUiType("crawl_liker.ui")[0]
class INSTA_Window(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.btn_start_recent.clicked.connect(self.btn_start_recent_clicked)  # 최근 게시물 크롤링
        self.btn_start_target.clicked.connect(self.btn_start_target_clicked)  # 특정 게시물 크롤링
        self.scroll_bar = self.log_browser.verticalScrollBar()
        self.log_browser.append(" - 추출 결과는 output 폴더 하위에 저장됩니다.")
        self.log_browser.append(" - 게시글 URL 예시) https://www.instagram.com/p/Cf-vUAcLS8K/")
        self.log_browser.append(" - 과도한 크롤링은 인스타그램에서 제한될 수 있습니다.")
        self.browser = execute_browser()
        self.liker_worker = LikerWorker(self)
        self.login_id = None
        self.login_pw = None
        self.file_name = None
        self.crawl_type = None

    # 최근 게시물 클릭
    def btn_start_recent_clicked(self):
        self.info("최근 게시물 좋아요 추출을 시작합니다.")
        self.button_activate(False)
        self.set_id_pw()
        if self.login():
            self.set_file_name('recent')
            create_excel(self.file_name)
            self.crawl_type = 'recent'
            self.liker_worker.start()
        else:
            self.button_activate(True)

    # 특정 게시물 클릭
    def btn_start_target_clicked(self):
        self.info("특정 게시물 좋아요 추출을 시작합니다.")
        self.button_activate(False)
        self.set_id_pw()
        if self.login():
            self.set_file_name('target')
            create_excel(self.file_name)
            self.crawl_type = 'target'
            self.liker_worker.start()
        else:
            self.button_activate(True)

    # 로그인 처리
    def login(self):
        if self.login_id is None or self.login_pw is None \
                or len(self.login_id) == 0 or len(self.login_pw) == 0:
            self.error("인스타그램 ID와 PW를 입력해주세요.")
            return False

        self.browser.get(Config.LOGIN_URL)
        sleep(2)

        if 'accounts' in self.browser.current_url:
            inputs = self.browser.find_elements(by=By.TAG_NAME, value='input')
            inputs[0].clear()
            inputs[1].clear()
            inputs[0].send_keys(self.login_id)
            inputs[1].send_keys(self.login_pw)
            inputs[1].submit()
            sleep(2)

            if '잘못된 비밀번호' in self.browser.page_source:
                self.error('잘못된 비밀번호입니다.')
                return False
            elif '입력한 사용자 이름' in self.browser.page_source:
                self.error('잘못된 사용자 ID입니다.')
                return False
            elif '문제가 발생' in self.browser.page_source:
                self.error('일시적인 문제가 발생하였습니다.')
                return False
            else:
                self.info(f'로그인에 성공했습니다. ID : {self.login_id}')
                return True

        else:  # 이미 로그인된 상태
            return True

    # ID/PW 세팅
    def set_id_pw(self):
        self.login_id = self.edit_id.text()
        self.login_pw = self.edit_pw.text()

    # 파일명 세팅
    def set_file_name(self, crawl_type):
        input_file_name = self.edit_file_name.text()
        if len(input_file_name) > 0:
            file_name = input_file_name + '.xlsx'
        else:
            now = datetime.now().strftime('%Y%m%d%H%M%S')
            type_name = '최근' if crawl_type == 'recent' else '특정'
            file_name = f'{now}_{type_name} 게시물 좋아요 추출.xlsx'

        self.file_name = file_name

    # 버튼 비활성화
    def button_activate(self, enable):
        self.btn_start_recent.setEnabled(enable)
        self.btn_start_target.setEnabled(enable)

    # 디버그 로그 출력
    def debug(self, text):
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f'<p>[{now}] {text}</p>')
        self.scroll_bar.setValue(self.scroll_bar.maximum())

    # 정보 로그 출력(초록색)
    def info(self, text):
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f'<p style="color:green">[{now}] {text}</p>')
        self.scroll_bar.setValue(self.scroll_bar.maximum())

    # 에러 로그 출력(빨간색)
    def error(self, text):
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f'<p style="color:red">[{now}] {text}</p>')
        self.scroll_bar.setValue(self.scroll_bar.maximum())

    # 종료 이벤트
    def closeEvent(self, QCloseEvent):
        if hasattr(self, 'browser') and self.browser is not None:
            self.browser.quit()

def create_excel(file_name):
    wb = Workbook()
    wb.create_sheet('게시글1', 0)
    wb.create_sheet('게시글2', 1)
    wb.create_sheet('게시글3', 2)
    wb.create_sheet('게시글4', 3)
    wb.create_sheet('게시글5', 4)
    wb.create_sheet('게시글6', 5)
    wb.create_sheet('공통 사용자', 6)

    ws1 = wb['게시글1']
    ws2 = wb['게시글2']
    ws3 = wb['게시글3']
    ws4 = wb['게시글4']
    ws5 = wb['게시글5']
    ws6 = wb['게시글6']
    ws7 = wb['공통 사용자']

    sub_liker = ['구분', '번호', '사용자ID', '사용자명']
    for kwd, j in zip(sub_liker, list(range(1, len(sub_liker) + 1))):
        ws1.cell(row=1, column=j).value = kwd
        ws2.cell(row=1, column=j).value = kwd
        ws3.cell(row=1, column=j).value = kwd
        ws4.cell(row=1, column=j).value = kwd
        ws5.cell(row=1, column=j).value = kwd
        ws6.cell(row=1, column=j).value = kwd

    sub_total = ['번호', '사용자ID']
    for kwd, j in zip(sub_total, list(range(1, len(sub_total) + 1))):
        ws7.cell(row=1, column=j).value = kwd

    file_path = Config.FILE_PATH
    if not os.path.isdir(file_path):
        os.mkdir(file_path)
    wb.save(file_path + file_name)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    INSTA = INSTA_Window()
    INSTA.show()
    app.exec()
