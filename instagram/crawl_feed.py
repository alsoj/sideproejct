import os
import sys

from PyQt6.QtWidgets import *
from PyQt6.QtCore import QDate
from PyQt6 import uic

from datetime import datetime, date, timedelta
from Common import execute_browser, debug, info, error, is_null, get_hashtag, download_img
from openpyxl import Workbook, load_workbook

import Config

from Login import LoginWorker
from Profile import ProfileWorker
from Timeline import TimelineWorker

form_class = uic.loadUiType("crawl_feed.ui")[0]
class INSTA_Window(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        # 메인 함수 세팅
        self.btn_start.clicked.connect(self.btn_start_clicked)

        # 로그 브라우저 세팅
        self.log_browser.append(" - 추출 결과는 output 폴더 하위에 저장됩니다.")
        self.log_browser.append(" - '대상ID_추출일시'로 폴더가 자동 생성됩니다.")
        self.log_browser.append(" - 과도한 크롤링은 인스타그램에서 제한될 수 있습니다.")
        self.log_browser.append(" - 사용 중 문의사항은 카카오톡 @lazymarkter로 연락주세요.")

        # 날짜 기본 세팅
        self.edit_date_from.setDate(QDate(datetime.now()-timedelta(days=30)))
        self.edit_date_to.setDate(QDate(datetime.now()))
        self.check_date_all.stateChanged.connect(self.check_all_period)

        # 테스트 정보 세팅
        # self.edit_login_id.setText('lazybrothers@naver.com')
        # self.edit_login_pw.setText('Lazy2210!')
        # self.edit_target_id.setText('go.ownos')
        # self.edit_target_id.setText('handok_motors')

        self.browser = execute_browser()
        self.login_id = None
        self.login_pw = None
        self.target_id = None
        self.traget_user_profile = None
        self.all_period = False

        self.login_worker = LoginWorker(self)
        self.profile_worker = ProfileWorker(self)
        self.timeline_worker = TimelineWorker(self)

        self.directory_name = None
        self.file_name = None

    # 피드 크롤링 클릭
    def btn_start_clicked(self):

        # 버튼 비활성화
        self.button_activate(False)

        # 추출 대상 체크
        target_id = self.edit_target_id.text()
        if is_null(target_id):
            self.button_activate(True)
            error(self.log_browser, "추출 대상 ID를 입력해주세요.")
            return

        if self.edit_date_to.date().toPyDate() < self.edit_date_from.date().toPyDate():
            error(self.log_browser, "종료일자는 시작일자보다 크거나 같아야 합니다.")
            return

        # 디렉토리 생성
        now = datetime.now().strftime("%Y%m%d%H%M%S")
        self.directory_name = f'{target_id}_{now}/'
        self.file_name = f'{target_id}_{now}.xlsx'
        create_directory(self.directory_name)
        create_excel(self.directory_name, self.file_name)

        # 로그인
        self.set_id_pw()
        self.login_worker.start()

    # ID/PW 세팅
    def set_id_pw(self):
        self.login_id = self.edit_login_id.text()
        self.login_pw = self.edit_login_pw.text()

    # 로그인 콜백
    def after_login(self):
        info(self.log_browser, "사용자 피드 추출을 시작합니다.")

        # 사용자 프로필 추출
        self.profile_worker.set_target_id(self.edit_target_id.text())
        self.profile_worker.run()

    # 사용자 정보 추출
    def after_profile(self):
        user_profile = self.profile_worker.user_profile

        wb = load_workbook(Config.FILE_PATH + self.directory_name + self.file_name, data_only=True)
        ws = wb.active
        ws.append([''])
        ws.append(['','사용자 ID', user_profile['username']])
        ws.append(['','사용자 명', user_profile['full_name']])
        ws.append(['','계정 설명', user_profile['biography']])
        ws.append(['','게시글 수', user_profile['edge_owner_to_timeline_media']['count']])
        ws.append(['','팔로워 수', user_profile['edge_followed_by']['count']])
        ws.append(['','팔로우 수', user_profile['edge_follow']['count']])
        ws.append([''])
        ws.append(['','작성자ID', '작성일시', '게시글 URL', '좋아요 수', '댓글 수', '내용', '태그'])
        wb.save(Config.FILE_PATH + self.directory_name + self.file_name)

        # 사용자 피드 추출
        self.timeline_worker.set_target_id(self.edit_target_id.text())
        if self.all_period:
            from_date = date(1, 1, 1)
            to_date = date(9999, 12, 31)
        else:
            from_date = self.edit_date_from.date().toPyDate()
            to_date = self.edit_date_to.date().toPyDate()
        self.timeline_worker.set_target_date_from(from_date)
        self.timeline_worker.set_target_date_to(to_date)
        self.timeline_worker.run()

    # 사용자 피드 추출 콜백
    def after_timeline(self):
        wb = load_workbook(Config.FILE_PATH + self.directory_name + self.file_name, data_only=True)
        ws = wb.active

        for result in self.timeline_worker.result_list:
            ws.append([
                '',
                result['username'],
                result['taken_at'],
                f"https://www.instagram.com/p/{result['code']}/",
                result['like_count'],
                result['comment_count'],
                result['caption'],
                ' \r\n'.join(str(s) for s in get_hashtag(result['caption']))
            ])

            # 이미지 다운로드
            download_img(result['image_url'], Config.FILE_PATH + self.directory_name + result['code'])
            wb.save(Config.FILE_PATH + self.directory_name + self.file_name)

        wb.save(Config.FILE_PATH + self.directory_name + self.file_name)
        self.button_activate(True)
        info(self.log_browser, "사용자 피드 추출이 완료되었습니다.")

    # 전체 기간 선택
    def check_all_period(self, state):
        enabled = state == 0  # 2: 체크, 0:체크해제
        self.edit_date_from.setEnabled(enabled)
        self.edit_date_to.setEnabled(enabled)
        self.all_period = not enabled

    # 버튼 비활성화
    def button_activate(self, enable):
        self.btn_start.setEnabled(enable)
        self.edit_login_id.setEnabled(enable)
        self.edit_login_pw.setEnabled(enable)
        self.edit_target_id.setEnabled(enable)

        if enable and self.all_period:  # 활성화 시에는 전체기간 선택 상태면 기간 선택만 활성화
            self.check_date_all.setEnabled(enable)
        else:
            self.check_date_all.setEnabled(enable)
            self.edit_date_from.setEnabled(enable)
            self.edit_date_to.setEnabled(enable)

    # 종료 이벤트
    def closeEvent(self, QCloseEvent):
        if hasattr(self, 'browser') and self.browser is not None:
            self.browser.quit()

# 디렉토리 생성
def create_directory(directory_name):
    file_path = Config.FILE_PATH
    if not os.path.isdir(file_path + directory_name):
        os.mkdir(file_path + directory_name)

# 엑셀 파일 생성
def create_excel(directory_name, file_name):
    wb = Workbook()
    wb.save(Config.FILE_PATH + directory_name + file_name)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    INSTA = INSTA_Window()
    INSTA.show()
    app.exec()
