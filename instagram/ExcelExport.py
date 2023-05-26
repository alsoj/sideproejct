from openpyxl import Workbook, load_workbook
import Config
import os
import datetime

FILE_NAME = ''
FILE_SUFFIX = '.xlsx'

def get_new_filename(crawl_type, add_param=None):
    filename = ''
    NOW = datetime.datetime.now().strftime('%Y%m%d%H%M%S')

    if crawl_type == 'follower':
        filename = f'{NOW}_팔로워 추출({add_param}).xlsx'
    elif crawl_type == 'following':
        filename = f'{NOW}_팔로잉 추출({add_param}).xlsx'

    return filename

def create_excel(crawl_type, file_name):
    wb = Workbook()
    ws = wb.active

    title = ''
    if crawl_type == 'like':
        title = ['번호', '사용자ID', '사용자명']
    elif crawl_type == 'comment':
        title = ['코드', '번호', '레벨', '사용자ID', '내용', '작성일시']
    elif crawl_type == 'timeline':
        title = ['사용자ID', '게시일시', '코드', '좋아요 수', '댓글 수']
    elif crawl_type == 'tag':
        title = ['번호', '사용자ID', '사용자명', '팔로워 수', '팔로잉 수']
    elif crawl_type == 'follow':
        title = ['추출 ID', '사용자ID', '사용자명']
    elif crawl_type == 'post':
        title = ['사용자ID', '사용자명', '게시일시', '코드', '좋아요 수', '댓글 수', '내용', '장소', '태그유저']
    elif crawl_type == 'user':
        title = ['사용자ID', '사용자명', '게시글', '팔로워', '팔로잉']

    for kwd, j in zip(title, list(range(1, len(title) + 1))):
        ws.cell(row=1, column=j).value = kwd

    file_path = Config.FILE_PATH
    if not os.path.isdir(file_path):
        os.mkdir(file_path)
    wb.save(file_path + file_name)


# 엑셀 입력
def write_excel(row_list, filename):
    file_path_name = Config.FILE_PATH + filename
    wb = load_workbook(file_path_name, data_only=True)
    ws = wb.active

    for row in row_list:
        ws.append(row)
    wb.save(file_path_name)

def get_workbook(filename):
    file_path_name = Config.FILE_PATH + filename
    wb = load_workbook(file_path_name, data_only=True)
    return wb

def save_workbook(wb, filename):
    file_path_name = Config.FILE_PATH + filename
    wb.save(file_path_name)