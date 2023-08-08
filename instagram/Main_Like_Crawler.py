import os
import sys
from time import sleep

from PyQt6.QtWidgets import *
from PyQt6 import uic

from datetime import datetime
from Common import execute_browser
from openpyxl import Workbook

import Config
import Common
from Liker import LikerWorker
from selenium.webdriver.common.by import By

from instagram.Login import LoginWorker

form_class = uic.loadUiType("Main_Like_Crawler.ui")[0]

class INSTA_Window(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.btn_start_recent.clicked.connect(self.btn_start_recent_clicked)  # 최근 게시물 크롤링
        self.scroll_bar = self.log_browser.verticalScrollBar()
        self.log_browser.append(" - 추출 결과는 output 폴더 하위에 저장됩니다.")
        self.log_browser.append(" - 게시글 URL 예시) https://www.instagram.com/p/Cf-vUAcLS8K/")
        self.log_browser.append(" - 과도한 크롤링은 인스타그램에서 제한될 수 있습니다.")
        self.browser = execute_browser(background=False)
        self.liker_worker = LikerWorker(self)
        self.login_worker = LoginWorker(self)
        self.login_id = None
        self.login_pw = None
        self.file_name = None
        self.target_id = None
        self.target_count = None
        self.crawl_type = 'recent'

        # Test용
        self.edit_id.setText("lazybrothers@naver.com")
        self.edit_pw.setText("Lazy2210!")

    # 최근 게시물 클릭
    def btn_start_recent_clicked(self):
        Common.info(self.log_browser, "최근 게시물 좋아요 추출을 시작합니다.")
        self.button_activate(False)
        self.set_id_pw()
        self.login_worker.start()

    # ID/PW 세팅
    def set_id_pw(self):
        self.login_id = self.edit_id.text()
        self.login_pw = self.edit_pw.text()

    def after_login(self):
        self.target_id = self.edit_target_id.text()
        self.file_name = get_file_name(self.target_id)
        self.target_count = self.spin_target_count.value()

        create_excel(self.file_name, self.target_count)
        self.liker_worker.start()

    # 버튼 비활성화
    def button_activate(self, enable):
        self.btn_start_recent.setEnabled(enable)

    # 종료 이벤트
    def closeEvent(self, QCloseEvent):
        if hasattr(self, 'browser') and self.browser is not None:
            self.browser.quit()

# 파일명 세팅
def get_file_name(target_id):
    now = datetime.now().strftime('%Y%m%d%H%M%S')
    file_name = f'{now}_좋아요 추출({target_id}).xlsx'
    return file_name

# 엑셀 파일 세팅
def create_excel(file_name, target_count):
    wb = Workbook()
    sub_liker = ['구분', '번호', '사용자ID', '사용자명']

    for i in range(target_count):
        wb.create_sheet('게시글' + str(i + 1), i)

        for kwd, j in zip(sub_liker, list(range(1, len(sub_liker) + 1))):
            wb['게시글' + str(i + 1)].cell(row=1, column=j).value = kwd

    file_path = Config.FILE_PATH
    if not os.path.isdir(file_path):
        os.mkdir(file_path)
    wb.save(file_path + file_name)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    INSTA = INSTA_Window()
    INSTA.show()
    app.exec()
