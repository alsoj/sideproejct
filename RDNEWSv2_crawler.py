import sys
import os
import platform

from PyQt6.QtCore import QDate, QThread
from PyQt6.QtWidgets import *

from PyQt6.QtWidgets import QMainWindow
from datetime import datetime, timedelta
from PyQt6 import uic, QtGui, QtCore

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By

from openpyxl import Workbook, load_workbook

from time import sleep
import unicodedata
import re

# import logging
# logger = logging.getLogger()  # 로그 생성
# logger.setLevel(logging.DEBUG)  # 로그의 출력 기준 설정
# formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')  # log 출력 형식
#
# # log를 파일에 출력
# file_handler = logging.FileHandler('RDNEWS_crawler.log')
# file_handler.setFormatter(formatter)
# logger.addHandler(file_handler)

# 전역변수 - 파일 관련
FILE_PATH = './output/'
FILE_PREFIX = 'RDNEWS_'
FILE_SUFFIX = '.xlsx'
FILE_NAME = ''

# 전역변수 - 재시도 관련
SLEEP_TIME = 60
RETRY_CNT = 10

form_class = uic.loadUiType("RDNEWSv2_crawler.ui")[0]
class RDNEWSCrawler(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.news_worker = NewsWorker(self)

        # 날짜 기본 세팅
        self.input_start_ymd.setDate(QDate(datetime.now()-timedelta(days=7)))
        self.input_end_ymd.setDate(QDate(datetime.now()))
        self.scroll_bar = self.log_browser.verticalScrollBar()

        # 작업 시작 클릭 시
        self.btn_start.clicked.connect(self.crawl_news)

    def debug(self, text):  # 디버그 로그 출력
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f"[{now}] " + text)
        self.scroll_bar.setValue(self.scroll_bar.maximum())

    def info(self, text):  # 정보 로그 출력
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f'<p style="color:green">[{now}] {text}</p>')
        self.scroll_bar.setValue(self.scroll_bar.maximum())

    def error(self, text):  # 에러 로그 출력
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_browser.append(f'<p style="color:red">[{now}] {text}</p>')
        self.scroll_bar.setValue(self.scroll_bar.maximum())

    def crawl_news(self):
        if self.input_end_ymd.date().toPyDate() < self.input_start_ymd.date().toPyDate():
            self.error("오류 발생 : 종료일자는 시작일자보다 크거나 같아야 합니다.")
        else:
            background_yn = "Y" if self.chk_background.isChecked() > 0 else "N"
            options = get_option(background_yn)
            self.browser = get_browser(options)

            if self.browser is not None:
                self.news_worker.start()
            else:
                self.error("크롬 브라우저 실행에 실패했습니다.")

    def closeEvent(self, QCloseEvent):
        if hasattr(self, 'browser'):
            self.browser.quit()

class NewsWorker(QThread):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        try:
            self.parent.btn_start.setEnabled(False)
            self.parent.info("노동신문 크롤링 START")

            start_ymd = self.parent.input_start_ymd.date().toPyDate()
            end_ymd = self.parent.input_end_ymd.date().toPyDate()
            str_start_ymd = start_ymd.strftime("%Y%m%d")
            str_end_ymd = end_ymd.strftime("%Y%m%d")

            # 엑셀 세팅
            global FILE_NAME
            FILE_NAME = FILE_PREFIX + str_start_ymd + "_" + str_end_ymd + FILE_SUFFIX

            create_excel()
            wb = load_workbook(FILE_PATH + FILE_NAME, data_only=True)
            ws = wb.active

            # INDEX 페이지 이동
            index_url = 'http://www.rodong.rep.kp/ko'

            crawl_ymd = start_ymd
            while crawl_ymd <= end_ymd:
                # 대상 날짜 이동
                crawl_input_ymd = crawl_ymd.strftime("%Y%m%d")
                str_input_ymd = crawl_ymd.strftime("%Y-%m-%d")
                self.parent.info(f"{str_input_ymd}의 기사 수집 중")
                self.connect_url(index_url)
                set_date(self.parent.browser, crawl_input_ymd)

                url_list, title_list = get_url_list(self.parent.browser)

                for i, url in enumerate(url_list):
                    if self.connect_url(url):  # 연결되었을 때 상세 크롤링 진행
                        # 기사 상세 크롤링
                        temp_row = get_detail(self.parent.browser)
                        key = str_input_ymd + "-" + str(i+1).zfill(3)
                        temp_row.insert(0, key)  # 일련번호
                        temp_row.insert(4, str_input_ymd)  # 일자
                        temp_row.insert(5, "《로동신문》")  # 매체명
                        news_side = re.findall('\d면', title_list[i])
                        if len(news_side) > 0:
                            temp_row.insert(6, news_side[0])  # 면수
                        else:
                            temp_row.insert(6, '')  # 면수

                        # 엑셀 append
                        ws.append(temp_row)
                        wb.save(FILE_PATH + FILE_NAME)
                        self.parent.debug(f"{str_input_ymd} : {str(i+1)}번째 기사 수집 중")

                crawl_ymd += timedelta(days=1)

            self.parent.info("노동신문 크롤링 END")
        except Exception as e:
            self.parent.error(f"크롤링 진행 중 오류 발생 : {str(e)}")
        finally:
            self.parent.browser.quit()
            self.parent.btn_start.setEnabled(True)

    def connect_url(self, url):
        retries = RETRY_CNT
        connected = False
        while retries > 0:
            try:
                self.parent.browser.get(url)
                connected = True
                break
            except Exception as e:
                retries -= 1
                self.parent.debug("응답이 없어 재시도 합니다. 남은 재시도 회수 : " + str(retries))
                sleep(SLEEP_TIME)
                continue
        return connected

def get_browser(options):
    browser = None
    try:
        browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    except Exception as e:
        browser = webdriver.Chrome(executable_path='./chromedriver', options=options)
    finally:
        return browser

def get_option(background_yn):
    # 백그라운드 실행 세팅
    options = webdriver.ChromeOptions()
    if background_yn == "Y":
        options.add_argument("headless")
    return options

# 해당 연도 선택
def select_year(browser, input_year):
    cal = browser.find_element(by=By.ID, value='topDate')
    cal.click()
    year = cal.find_element(by=By.CLASS_NAME, value='dhtmlxcalendar_month_label_year')
    year.click()

    table = cal.find_element(by=By.CLASS_NAME, value='dhtmlxcalendar_area_selector_year')
    lis = table.find_elements(by=By.TAG_NAME, value='li')

    left = cal.find_element(by=By.CLASS_NAME, value='dhtmlxcalendar_selector_cell_left')
    right = cal.find_element(by=By.CLASS_NAME, value='dhtmlxcalendar_selector_cell_right')

    # 해당 연도가 있는 page로 이동
    target_found = False
    while not target_found:
        year_set = set()
        for li in lis:
            year_set.add(li.text)

        if max(year_set) < input_year:
            right.click()
        elif min(year_set) > input_year:
            left.click()
        else:
            target_found = True

    # 해당 연도 클릭
    for li in lis:
        if li.text.strip() == input_year:
            li.click()
            break

# 해당 월 선택
def select_month(browser, input_month):
    cal = browser.find_element(by=By.ID, value='topDate')
    month = cal.find_element(by=By.CLASS_NAME, value='dhtmlxcalendar_month_label_month')
    month.click()

    table = cal.find_element(by=By.CLASS_NAME, value='dhtmlxcalendar_area_selector_month')
    lis = table.find_elements(by=By.TAG_NAME, value='li')

    # 해당 월 클릭
    for li in lis:
        if li.text.strip().replace("월", "").zfill(2) == input_month:
            li.click()
            break

# 해당 일 선택
def select_day(browser, input_day):
    cal = browser.find_element(by=By.ID, value='topDate')
    dates = cal.find_element(by=By.CLASS_NAME, value='dhtmlxcalendar_dates_cont')
    lis = dates.find_elements(by=By.TAG_NAME, value='li')

    for li in lis:
        if 'dhtmlxcalendar_cell_month' in li.get_attribute('class') and li.text.strip().zfill(2) == input_day:
            li.click()
            break

# 날짜 선택
def set_date(browser, input_date):
    input_year = input_date[0:4]
    input_month = input_date[4:6]
    input_day = input_date[6:8]

    select_year(browser, input_year)
    select_month(browser, input_month)
    select_day(browser, input_day)

def get_url_list(browser):
    retries = RETRY_CNT
    url_list, title_list = [], []
    while retries > 0:
        try:
            newslist = browser.find_element(by=By.CLASS_NAME, value='date_news_list')
            rows = newslist.find_elements(by=By.CLASS_NAME, value='row')
            for row in rows:
                media_body = row.find_element(by=By.CLASS_NAME, value='media-body')
                a_tag = media_body.find_element(by=By.TAG_NAME, value='a')
                url_list.append(a_tag.get_attribute('href'))
                title_list.append(a_tag.text)
            break
        except Exception as e:
            retries -= 1
            print("get_url_list 오류 발생" + str(retries))
            sleep(5)
            continue
    return url_list, title_list

# 기사 상세 크롤링
def get_detail(browser):
    retries = RETRY_CNT
    temp_row = []
    while retries > 0:
        try:
            # 헤드
            result_head = ''
            heads = browser.find_elements(by=By.CLASS_NAME, value='news_Head')
            for head in heads:
                result_head += head.text + "\r\n"
            result_head = result_head.strip()

            # 제목
            result_title = ''
            titles = browser.find_elements(by=By.CLASS_NAME, value='news_Title')
            for title in titles:
                result_title += title.text + "\r\n"
            result_title = result_title.strip()

            # 부제
            result_subtitle = ''
            subtitles = browser.find_elements(by=By.CLASS_NAME, value='news_SubTitle')
            for subtitle in subtitles:
                result_subtitle += subtitle.text + "\r\n"
            result_subtitle = result_subtitle.strip()

            # 게재 일자
            # result_news_date = browser.find_element(by=By.CLASS_NAME, value='NewsDate').text

            # 신문 명, 면수
            # news_side = browser.find_element(by=By.CLASS_NAME, value='NewsSide').text
            # result_news_name = news_side.split(" ")[0]
            # result_news_side = news_side.split(" ")[1]

            # 기사내용, 기자 명
            result_content, result_writer = '', ''
            article_contents = browser.find_elements(by=By.CLASS_NAME, value='ArticleContent')
            for content in article_contents:
                if 'right' in content.get_attribute('style'):
                    result_writer += content.text
                else:
                    result_content += content.text + "\r\n"
            result_content = result_content.strip()

            temp_row.append(unicodedata.normalize('NFKC', result_head))
            temp_row.append(unicodedata.normalize('NFKC', result_title))
            temp_row.append(unicodedata.normalize('NFKC', result_subtitle))
            # temp_row.append(unicodedata.normalize('NFKC', result_news_date))
            # temp_row.append(unicodedata.normalize('NFKC', result_news_name))
            # temp_row.append(unicodedata.normalize('NFKC', result_news_side))
            temp_row.append(unicodedata.normalize('NFKC', result_content))
            temp_row.append(unicodedata.normalize('NFKC', result_writer))
            break
        except Exception as e:
            videos = browser.find_elements(by=By.TAG_NAME, value='video')
            if len(videos) > 0:
                retries = 0
            else:
                retries -= 1
                sleep(5)
                print("응답이 없어 재시도 합니다. 남은 재시도 회수 : ", retries)
                continue
    return temp_row

# 기사 상세 다음 페이지 이동
def go_to_next(browser):
    next_button = browser.find_element(by=By.ID, value='nextNews')
    next_button.click()

def create_excel():
    # 엑셀 생성
    wb = Workbook()
    ws = wb.active

    # 제목 적기
    sub = ['일련번호', '헤드', '기사 제목', '부제', '게재 일자', '신문 명', '면수', '기사내용', '기자 명']
    for kwd, j in zip(sub, list(range(1, len(sub) + 1))):
      ws.cell(row=1, column=j).value = kwd

    if not os.path.isdir(FILE_PATH):
        os.mkdir(FILE_PATH)

    wb.save(FILE_PATH+FILE_NAME)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    news_crawler = RDNEWSCrawler()
    news_crawler.show()
    app.exec()