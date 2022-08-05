import sys
from PyQt6.QtWidgets import *
from PyQt6 import uic

# 어플리케이션 패키지
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By

from openpyxl import Workbook, load_workbook
import datetime
from time import sleep
import unicodedata
import re

# 전역변수 - 파일 관련
FILE_PATH = './output/'
FILE_PREFIX = 'RDNEWS_'
FILE_SUFFIX = '.xlsx'
FILE_NAME = ''

# 전역변수 - 재시도 관련
SLEEP_TIME = 60
RETRY_CNT = 10

# 전역변수 - 백그라운드 실행 관련
BACKGROUND_YN = 'Y'

form_class = uic.loadUiType("RDNEWS_crawler.ui")[0]

class MyWindow(QMainWindow, form_class):
  def __init__(self):
    super().__init__()
    self.setupUi(self)

    self.btn_start.clicked.connect(self.btn_start_clicked)

  def btn_start_clicked(self):
    try:
      global BACKGROUND_YN
      BACKGROUND_YN = "Y" if self.chk_background.isChecked() > 0 else "N"

      self.log_text_browser.append(" ")
      self.log_text_browser.append("#############################################")
      self.log_text_browser.append("크롤링 작업 시작, 백그라운드 실행 : " + BACKGROUND_YN)
      self.log_text_browser.append("#############################################")
      QApplication.processEvents()

      start_ymd = self.input_start_ymd.date().toPyDate()
      end_ymd = self.input_end_ymd.date().toPyDate()
      str_start_ymd = start_ymd.strftime("%Y%m%d")
      str_end_ymd = end_ymd.strftime("%Y%m%d")
      self.log_text_browser.append("시작일자 :" +  str_start_ymd +  " / 종료일자 :" +  str_end_ymd)
      QApplication.processEvents()

    except Exception as e:
      print(e)

    if check_input_valid(start_ymd, end_ymd) :
      global FILE_NAME
      FILE_NAME = FILE_PREFIX + str_start_ymd + '_' + str_end_ymd + FILE_SUFFIX
      self.log_text_browser.append("생성 파일명 : " + FILE_NAME)
      QApplication.processEvents()

      create_excel()  # 엑셀파일 생성

      self.log_text_browser.append(" ")
      self.log_text_browser.append("크롬 브라우저가 실행 중입니다.")
      QApplication.processEvents()

      # 백그라운드 실행 세팅
      options = webdriver.ChromeOptions()
      if BACKGROUND_YN == "Y":
        options.add_argument("headless")

      browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

      try :
        crawl_ymd = start_ymd

        while (crawl_ymd <= end_ymd):
          crawl_input_ymd = crawl_ymd.strftime("%Y-%m-%d")
          crawl_by_date(self, browser, crawl_input_ymd)
          crawl_ymd += datetime.timedelta(days=1)

      except Exception as e:
        self.log_text_browser.append(" ")
        self.log_text_browser.append("#############################################")
        self.log_text_browser.append("크롤링 작업 오류 발생")
        self.log_text_browser.append(str(e))
        self.log_text_browser.append("#############################################")
        QApplication.processEvents()

      finally:
        self.log_text_browser.append(" ")
        self.log_text_browser.append("#############################################")
        self.log_text_browser.append("크롤링 작업 종료")
        self.log_text_browser.append("#############################################")
        browser.quit()
        QApplication.processEvents()

    else :
      self.log_text_browser.append("오류 발생 : 종료일자는 시작일자보다 크거나 같아야 합니다.")


####################################################
# 입력된 날짜 값이 정상인지 확인(시작일자 <= 종료일자)
####################################################
def check_input_valid(start_ymd, end_ymd) :
  if end_ymd < start_ymd :
    return False
  else :
    return True

####################################################
# 엑셀 파일 생성
####################################################
def create_excel():
    # 엑셀 생성
    wb = Workbook()
    ws = wb.active

    # 제목 적기
    sub = ['식별키', '목록-번호', '목록-구분', '목록-제목', '목록-면수', '목록-작성자', '목록-일자', '본문-일자(원본)', '본문-일자', '본문-내용', '본문-제목',
           '본문-부제(제목앞)', '본문-부제(제목뒤)']
    for kwd, j in zip(sub, list(range(1, len(sub) + 1))):
      ws.cell(row=1, column=j).value = kwd

    wb.save(FILE_PATH+FILE_NAME)

# text 타입 구분
def get_text_type(line, detail_title, detail_content):
  font_tag = line.find_elements(by=By.TAG_NAME, value='font')
  if len(font_tag) > 0 and 'font-size' in font_tag[0].get_attribute('style'):
    if 'bold' in font_tag[0].get_attribute('style'):  # bold인 것들 주제목
      return "TITLE"
    elif len(detail_title) > 0 and len(detail_content) == 0:  # 제목은 나오고, 내용은 나오지 않은 경우
      return "POST_TITLE"
    elif len(detail_title) == 0 and len(detail_content) == 0:  # 제목과 내용이 아직 나오지 않은 경우
      return "PRE_TITLE"
    else:
      return "CONTENT"
  else:
    return "CONTENT"

####################################################
# 날짜별 크롤링
####################################################
def crawl_by_date(self, browser, date):
  self.log_text_browser.append(" ")
  self.log_text_browser.append("#############################################")
  self.log_text_browser.append(date + " 크롤링 작업 시작")
  QApplication.processEvents()

  wb = load_workbook(FILE_PATH+FILE_NAME, data_only=True)
  ws = wb.active

  list_base_url = 'http://www.rodong.rep.kp/ko/index.php?strPageID=SF01_01_03&strDate='
  # browser.get(list_base_url+date)
  connect(browser, list_base_url+date)

  news_line_list = browser.find_elements(by=By.CLASS_NAME, value='ListNewsLineContainer')

  for news_line in news_line_list:
    temp_row = []

    line_no = news_line.find_element(by=By.CLASS_NAME, value='ListNewsLineNo').text.strip()
    self.log_text_browser.append(line_no.replace('.','').strip() + "번째 기사 진행 중")
    QApplication.processEvents()

    line_title = news_line.find_element(by=By.CLASS_NAME, value='ListNewsLineTitleW').text.strip()
    line_writer = news_line.find_element(by=By.CLASS_NAME, value='ListNewsLineWriter').text.strip()
    line_date = news_line.find_element(by=By.CLASS_NAME, value='ListNewsLineDate').text.strip()

    article_title, article_type, article_page = get_article_info(line_title)

    temp_row.append(unicodedata.normalize('NFKC', line_no))
    temp_row.append(unicodedata.normalize('NFKC', article_type))
    temp_row.append(unicodedata.normalize('NFKC', article_title))
    temp_row.append(unicodedata.normalize('NFKC', article_page))
    temp_row.append(unicodedata.normalize('NFKC', line_writer))
    temp_row.append(unicodedata.normalize('NFKC', line_date))

    # 상세창 팝업
    title_popup = news_line.find_element(by=By.CLASS_NAME, value='ListNewsLineTitleW')
    title_popup_a = title_popup.find_element(by=By.TAG_NAME, value='a')
    title_popup_a.click()
    browser.switch_to.window(browser.window_handles[1])
    connect(browser, browser.current_url)

    key = ''
    detail_date_origin = ''
    detail_date = ''
    detail_title = ''
    detail_sub_title_pre = ''
    detail_sub_title_next = ''
    detail_content = ''

    try:
      key = browser.current_url[-15:]
      detail_date_origin = browser.find_element(by=By.CLASS_NAME, value='ArticleMenuDate').text
      detail_date = key[:10]
      detail_content_list = browser.find_elements(by=By.CLASS_NAME, value='ArticleContent')

      for detail_line in detail_content_list:
        line_type = get_text_type(detail_line, detail_title, detail_content)
        # print(line_type, detail_line.text)
        if line_type == "TITLE":
          detail_title += detail_line.text + '\r\n'
        elif line_type == "PRE_TITLE":
          detail_sub_title_pre += detail_line.text + '\r\n'  # 부제목(앞)
        elif line_type == "POST_TITLE":
          detail_sub_title_next += detail_line.text + '\r\n'  # 부제목(뒤)
        elif line_type == "CONTENT" and line_writer not in detail_line.text:  # 마지막 작성자 제외
          detail_content += detail_line.text + '\r\n'  # 본문 내용
    except Exception as e:
      self.log_text_browser.append(line_no.replace('.','').strip() + "번째 기사 진행 중 오류 발생 PASS")
      QApplication.processEvents()

    finally:
      temp_row.insert(0, key)
      temp_row.append(detail_date_origin)
      temp_row.append(detail_date)
      temp_row.append(unicodedata.normalize('NFKC', detail_content.strip()))
      temp_row.append(unicodedata.normalize('NFKC', detail_title.strip()))
      temp_row.append(unicodedata.normalize('NFKC', detail_sub_title_pre.strip()))
      temp_row.append(unicodedata.normalize('NFKC', detail_sub_title_next.strip()))
      ws.append(temp_row)
      wb.save(FILE_PATH + FILE_NAME)

    # 팝업창 close 후 객체 전환
    browser.close()
    browser.switch_to.window(browser.window_handles[0])

  self.log_text_browser.append("#############################################")
  QApplication.processEvents()

####################################################
# 목록-기사 정보 추출
####################################################
def get_article_info(title_full):
  try:
    if title_full.startswith('['):
      article_type = title_full[title_full.find('[') + 1: title_full.find(']')].strip()
      article_page = re.findall('\d면', title_full)[0]
    else:
      article_type = ''
      article_page = re.findall('\d면', title_full)[0]

    article_title = title_full
    article_title = article_title.replace(article_type, "")
    article_title = article_title.replace(article_page, "")
    article_title = article_title.replace("[", "")
    article_title = article_title.replace("]", "")
  except Exception as e:
    article_title = title_full
    article_type = ''
    article_page = ''

  return article_title.strip(), article_type.strip(), article_page.strip()

####################################################
# connect 연결(재시도)
####################################################
def connect(browser, url):
  retries = RETRY_CNT
  while (retries > 0):
    try:
      browser.get(url)
      break
    except Exception as e:
      retries = retries - 1
      sleep(SLEEP_TIME)
      print("응답이 없어 재시도 합니다. 남은 재시도 회수 : ", retries)
      print(e)
      continue

if __name__ == "__main__":
  app = QApplication(sys.argv)
  myWindow = MyWindow()
  myWindow.show()
  app.exec()