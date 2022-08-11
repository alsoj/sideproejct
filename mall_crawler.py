import sys
from PyQt6.QtWidgets import *
from PyQt6 import uic

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By

import requests

from openpyxl import Workbook, load_workbook
import datetime

from time import sleep
import re

# 전역변수 - 파일 관련
FILE_PATH = './output/'
FILE_SUFFIX = '.xlsx'
FILE_NAME = ''

INPUT_FULL_FILE_NAME = ''
INPUT_FILE_NAME = ''

# 전역변수 - 키워드
KEYWORD = ''
KEYWORD_LIST = []
PAGE = 1

# 전역젼수
BASE_URL = 'https://ad.search.naver.com/search.naver?where=ad&query='
SEARCH_URL = 'https://ad.search.naver.com/search.naver?where=ad&sm=svc_nrs&query={}&referenceId=&pagingIndex={}'

form_class = uic.loadUiType("mall_crawler.ui")[0]

class MallCrawler(QMainWindow, form_class):
  def __init__(self):
    super().__init__()
    self.setupUi(self)

    self.btn_start.clicked.connect(self.btn_start_clicked)
    self.btn_file.clicked.connect(self.btn_file_clicked)

  def btn_file_clicked(self):
    file_name = QFileDialog.getOpenFileName(self, '파일 선택', './', 'Excel(*.xlsx)')

    if file_name[0]:  # 파일 선택
      file_name = file_name[0]
    else:  # 파일 미선택
      return False

    self.edit_file.setText(file_name)

    global INPUT_FULL_FILE_NAME
    INPUT_FULL_FILE_NAME = file_name
    global INPUT_FILE_NAME
    INPUT_FILE_NAME = file_name.split("/")[-1].replace(".xlsx", "")
    global KEYWORD_LIST
    KEYWORD_LIST = get_keyword_list(file_name)

  def btn_start_clicked(self):
    self.log_text_browser.clear()
    self.btn_start.setEnabled(False)
    QApplication.processEvents()

    global KEYWORD_LIST
    global KEYWORD
    KEYWORD = self.edit_keyword.text().strip()
    if len(KEYWORD) + len(KEYWORD_LIST) == 0:
      self.btn_start.setEnabled(True)
      self.log_text_browser.append("파일 선택 또는 키워드를 입력 해주세요.")
      QApplication.processEvents()
      return False

    # 파일 입력을 우선해서 설정
    if len(KEYWORD_LIST) > 0:
      crawl_type = 'FILE'
    else:
      crawl_type = 'TEXT'
      KEYWORD_LIST.append(KEYWORD)

    # 엑셀 파일 생성
    create_excel(crawl_type)

    try:
      # 백그라운드 실행
      options = webdriver.ChromeOptions()
      options.add_argument("headless")
      browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
      browser.set_page_load_timeout(30)

      self.progress_bar.setMaximum(len(KEYWORD_LIST))
      i = 0

      for KEYWORD in KEYWORD_LIST:
        self.log_text_browser.append("#############################################")
        self.log_text_browser.append("크롤링 작업 시작, 키워드 : " + KEYWORD)
        QApplication.processEvents()

        # print(SEARCH_URL.format(KEYWORD, PAGE))
        global PAGE
        PAGE = 1

        browser.get(SEARCH_URL.format(KEYWORD, PAGE))
        progress = 'start'

        while (progress != 'end'):
          go_mall_in_page(self, browser, KEYWORD)  # 쇼핑몰로 이동
          progress = go_to_next(browser, KEYWORD)  # 다음페이지로 이동

        self.log_text_browser.append("크롤링 작업 종료, 키워드 : " + KEYWORD)
        self.log_text_browser.append("#############################################")
        i += 1
        self.progress_bar.setValue(i)
        QApplication.processEvents()

    except Exception as e:
      self.log_text_browser.append("#############################################")
      self.log_text_browser.append("작업 중 오류가 발생했습니다. " + str(e))
      self.log_text_browser.append("#############################################")
      QApplication.processEvents()

    finally:
      browser.quit()
      self.btn_start.setEnabled(True)
      self.log_text_browser.append("#############################################")
      self.log_text_browser.append("작업이 종료 되었습니다.")
      self.log_text_browser.append("#############################################")
      QApplication.processEvents()


def get_keyword_list(file_name):
  wb = load_workbook(file_name, data_only=True)
  ws = wb.active

  keyword_list = []

  for row in ws.rows:
    keyword_list.append(row[0].value)

  return keyword_list

####################################################
# 엑셀 파일 생성
####################################################
def create_excel(crawl_type):
  # 파일 생성
  global FILE_NAME
  if crawl_type == 'FILE':
    FILE_NAME = datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '_' + INPUT_FILE_NAME + FILE_SUFFIX
  elif crawl_type == 'TEXT':
    FILE_NAME = datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '_' + KEYWORD + FILE_SUFFIX

  # 엑셀 생성
  wb = Workbook()
  ws = wb.active

  # 제목 적기
  sub = ['검색어', '쇼핑몰 명', 'URL', '대표자 명', 'EMAIL', '전화번호']
  for kwd, j in zip(sub, list(range(1, len(sub) + 1))):
    ws.cell(row=1, column=j).value = kwd

  wb.save(FILE_PATH+FILE_NAME)

def is_soho_url(url):
  if 'smartstore' in url \
          or 'ssg.com' in url \
          or 'coupang' in url \
          or '11st' in url \
          or 'balaan' in url \
          or 'gmarket' in url \
          or 'ssfshop' in url \
          or 'auction' in url \
          or 'tmon' in url \
          or 'SSG.COM' in url \
          or 'danawa' in url \
          or 'lotte' in url \
          or 'gsshop' in url \
          or 'cj' in url \
          or 'hmall' in url \
          or 'naver' in url \
          or 'musinsa' in url \
          or 'wemakeprice' in url:
    return False
  else:
    return True

####################################
# 해당 페이지 내에 존재하는 쇼핑몰 이동
####################################
def go_mall_in_page(self, browser, KEYWORD):
  url_list = browser.find_elements(by=By.CLASS_NAME, value='url_area')

  for url in url_list:
    try:
      url = url.text.split(' ')[0]
      if is_soho_url(url):
        # selenium에서 tab을 띄워서 확인하는데 시간이 오래 소요돼서, requests로 선확인
        res = requests.get(url)
        res_text = res.text
        if 'cafe24' in res_text or '카페24' in res_text:
          self.log_text_browser.append("카페24 업체 PASS : " + url.split('/')[2])
          QApplication.processEvents()
          pass
        elif 'makeshop' in res_text:
          self.log_text_browser.append("메이크샵 업체 PASS : " + url.split('/')[2])
          QApplication.processEvents()
          pass
        else:
          browser.switch_to.new_window('tab')
          browser.get(url)
          sleep(2)
          get_mall_info(self, browser, KEYWORD)
      else:
        self.log_text_browser.append("비대상 업체 PASS : " + url.split('/')[2])
        QApplication.processEvents()
        pass
    except Exception as e:
      pass
    finally:
      close_new_tabs(browser)


  # 썸네일 이미지가 있는 건들만 List로 추출(썸네일 없는 건은 광고)
  # mall_thumb_list = browser.find_elements(by=By.CLASS_NAME, value='ad_thumb')
  #
  # for mall_thumb in mall_thumb_list:
  #   try:
  #     mall_thumb.click()
  #     sleep(3)
  #     get_mall_info(self, browser, KEYWORD)
  #   except Exception as e:
  #     close_new_tabs(browser)
  #     pass

####################################
# 다음 페이지로 이동
####################################
def go_to_next(browser, KEYWORD):

  global PAGE
  PAGE += 1
  browser.get(SEARCH_URL.format(KEYWORD, PAGE))

  if '검색결과가 없습니다' in browser.page_source :
    return 'end'
  else:
    return 'move'

    # next_button = browser.find_elements(by=By.CLASS_NAME, value='next')
    # if next_button :
    #     next_button[0].click()
    #     print("페이지 이동")
    #     return "move"
    # else :
    #     print("마지막 페이지입니다")
    #     return "end"

####################################
# ceo 명 조회
####################################
def get_ceo_name(source):
  source = source.upper()
  
  # 추출을 방해하는 키워드 제거
  source = source.replace('대표이미지', '')
  source = source.replace('대표번호', '')
  source = source.replace('대표 이미지', '')
  source = source.replace('대표 번호', '')
  source = source.replace('OWNERID', '')
  source = source.replace('&NBSP;', '')
  source = source.replace('<STRONG>', '')
  source = source.replace('</STRONG>', '')

  ceo = ''

  if len(source[source.rfind('대표자'):source.rfind('대표자')+20]) > 0:
      ceo = source[source.rfind('대표자'):source.rfind('대표자')+20]
  elif len(source[source.rfind('대표이사'):source.rfind('대표이사')+20]) > 0:
      ceo = source[source.rfind('대표이사'):source.rfind('대표이사')+20]
  elif len(source[source.rfind('CEO'):source.rfind('CEO') + 20]) > 0:
    ceo = source[source.rfind('CEO'):source.rfind('CEO') + 20]
  elif len(source[source.rfind('C.E.O'):source.rfind('C.E.O') + 20]) > 0:
    ceo = source[source.rfind('C.E.O'):source.rfind('C.E.O') + 20]
  elif len(source[source.rfind('OWNER'):source.rfind('OWNER') + 20]) > 0:
    ceo = source[source.rfind('OWNER'):source.rfind('OWNER') + 20]
  elif len(source[source.rfind('대표'):source.rfind('대표') + 20]) > 0:
    ceo = source[source.rfind('대표'):source.rfind('대표') + 20]

  # 조회어 제거
  ceo = ceo.replace('CEO', ' ')
  ceo = ceo.replace('C.E.O', ' ')
  ceo = ceo.replace('OWNER', ' ')
  ceo = ceo.replace('대표자', ' ')
  ceo = ceo.replace('대표이사', ' ')
  ceo = ceo.replace('대표', ' ')

  # 태그제거
  ceo = re.compile('(<([^>]+)>)').sub('', ceo)
  ceo = re.compile('(<([^>]+))').sub('', ceo)

  ceo = ceo.replace(',', ' ')
  ceo = re.compile('[^ ㄱ-ㅣ가-힣+|A-Z+|a-z+]').sub('', ceo)
  ceo = re.compile('\|.+').sub('', ceo)
  ceo = re.compile('\ㅣ.+').sub('', ceo)

  rtn_ceo = ''
  for text in ceo.split(' '):
    if len(text.strip()) > 0:
      rtn_ceo += text + ' '

  return rtn_ceo.strip()

####################################
# email 조회
####################################
def get_email(source):
  email = ''
  emails = re.findall(r'[\w\.-]+@[\w\.-]+', source)

  if len(emails) > 0:
    email = emails[-1].lower()

  return email.strip()

####################################
# 전화번호 조회
####################################
def get_tel(source):
  tel = ''
  tels1 = re.findall(r'(?<!-|\d)0\d{1,3}-\d{3,4}-\d{4}(?!-|\d)', source) # 000-0000-0000 형태
  tels2 = re.findall(r'(?<!-|\d)\d{4}-\d{4}(?!-|\d)', source) # 0000-0000 형태

  tels = tels1 + tels2
  tels = list(set(tels))

  if len(tels) > 0:
    tel = ','.join(tels)

  # if len(tels) > 0:
  #   tel = tels[-1]
  # else:
  #   tels = re.findall(r'(?<!-|\d)\d{4}-\d{4}(?!-|\d)', source)
  #   if len(tels) > 0:
  #     tel = tels[-1]

  return tel.strip()

####################################
# mall 정보 조회
####################################
def get_mall_info(self, browser, KEYWORD):
  # browser.switch_to.window(window_name=browser.window_handles[-1])

  source = browser.page_source
  cur_url = browser.current_url

  if '카페24' in source or 'cafe24' in source:
    self.log_text_browser.append("카페24 호스팅 업체 PASS : " + cur_url.split('/')[2])
    # self.log_text_browser.append("===================================================================")
    QApplication.processEvents()
    # print('카페24 호스팅 업체 PASS : ', cur_url.split('/')[2])
    # print("===================================================================")
  elif 'smartstore' in cur_url:
    self.log_text_browser.append("스마트스토어 업체 PASS : " + cur_url.split('/')[2] + "/" + cur_url.split('/')[3].split('?')[0])
    # self.log_text_browser.append("===================================================================")
    QApplication.processEvents()
    # print('스마트스토어 업체 PASS : ', cur_url.split('/')[2] + "/" + cur_url.split('/')[3].split('?')[0])
    # print("===================================================================")
  elif 'makeshop' in source:
    self.log_text_browser.append("메이크샵 호스팅 업체 PASS : " + cur_url.split('/')[2])
    # self.log_text_browser.append("===================================================================")
    QApplication.processEvents()
    # print('메이크샵 호스팅 업체 PASS : ', cur_url.split('/')[2])
    # print("===================================================================")
  elif 'musinsa' in cur_url or 'lfmall' in cur_url or 'lotteon' in cur_url or 'danawa' in cur_url or 'navar' in cur_url or 'blog' in cur_url or 'ssg' in cur_url:
    # self.log_text_browser.append("===================================================================")
    self.log_text_browser.append("비대상 업체 PASS : " + cur_url.split('/')[2])
    QApplication.processEvents()
  else:
    title = browser.title
    site_url = cur_url.split('/')[2]

    # 대표 추출
    ceo = get_ceo_name(source)

    # 이메일 추출
    email = get_email(source)

    # 전화번호 추출
    tel = get_tel(source)

    wb = load_workbook(FILE_PATH + FILE_NAME, data_only=True)
    ws = wb.active

    # sub = ['검색어', '쇼핑몰 명', 'URL', '대표자 명', 'EMAIL', '전화번호']
    temp_row = [KEYWORD, title, site_url, ceo, email, tel]
    # print("대상 사이트 추출 : ", title)
    # print("사이트 url : ", site_url)
    # print("대표 : ", ceo)
    # print("이메일 주소 : ", email)
    # print("전화번호 : ", tel)
    # print("===================================================================")

    ws.append(temp_row)
    wb.save(FILE_PATH+FILE_NAME)

  # close_new_tabs(browser)

####################################
# 팝업창 종료
####################################
def close_new_tabs(browser):
  tabs = browser.window_handles
  while len(tabs) != 1 :
    browser.switch_to.window(tabs[1])
    browser.close()
    tabs = browser.window_handles
  browser.switch_to.window(tabs[0])

if __name__ == '__main__':
  app = QApplication(sys.argv)
  mallCrawler = MallCrawler()
  mallCrawler.show()
  app.exec()