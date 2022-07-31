import sys
from PyQt6.QtWidgets import *
from PyQt6 import uic

# 어플리케이션 패키지
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By

import datetime
from time import sleep
import json
from openpyxl import Workbook, load_workbook
import unicodedata

BASE_URL = 'https://www.instagram.com/'
LOGIN_URL = BASE_URL + 'accounts/login/'
LIKER_URL = BASE_URL + 'graphql/query/?query_hash=d5d763b1e2acf209d62d22d184488e57'
COMMENT_URL = BASE_URL + 'graphql/query/?query_hash=bc3296d1ce80a24b1b6e40b1e72903f5'

FILE_PATH = './output/'
FILE_NAME = ''
FILE_SUFFIX = '.xlsx'


form_class = uic.loadUiType("crawl_instagram.ui")[0]

class INSTA_Window(QMainWindow, form_class):
  def __init__(self):
    super().__init__()
    self.setupUi(self)
    self.btn_start.clicked.connect(self.btn_start_clicked)

  def btn_start_clicked(self):
    self.clearLog()
    self.setLogText("#######################################")
    self.setLogText("크롤링 작업을 시작합니다.")

    input_id = self.edit_id.text()
    input_pw = self.edit_pw.text()
    input_url = self.edit_url.text()

    if self.check_input_valid() is not True:
      return False

    options = webdriver.ChromeOptions()
    # options.add_argument("headless")
    browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    try:
      today = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
      short_code = get_short_code(input_url)
      if len(short_code) == 0:
        self.setLogText("게시물 URL이 올바르지 않습니다.")
        return False

      global FILE_NAME
      FILE_NAME = today + "_" + short_code + FILE_SUFFIX
      self.setLogText("생성 파일명 : " + FILE_NAME)
      create_excel()

      self.logBrowser.append("로그인 중 입니다.")
      login_instagram(browser, input_id, input_pw)
      sleep(5)

      self.logBrowser.append("좋아요 크롤링 중 입니다.")
      get_likers(browser, short_code)
      sleep(5)

      self.logBrowser.append("댓글 크롤링 중 입니다.")
      get_comments(browser, short_code)

    except Exception as e:
      self.setLogText("크롤링 작업 중 오류가 발생 했습니다.")
      self.setLogText(str(e))
      self.setLogText("#######################################")
    finally:
      browser.quit()
      self.setLogText("크롤링 작업이 종료되었습니다.")
      self.setLogText("#######################################")

  def check_input_valid(self):
    input_id = self.edit_id.text()
    input_pw = self.edit_pw.text()
    input_url = self.edit_url.text()

    if len(input_id) == 0 or len(input_pw) == 0 or len(input_url) == 0:
      self.setLogText("필수 입력 값이 입력되지 않았습니다.")
      self.setLogText("ID, PW, URL을 확인해주세요.")
      return False
    else:
      return True

  def clearLog(self):
    self.logBrowser.clear()

  def setLogText(self, text):
    self.logBrowser.append(text)
    QApplication.processEvents()

# 엑셀 파일 생성
def create_excel():
  wb = Workbook()
  wb.create_sheet('좋아요', 0)
  wb.create_sheet('댓글', 1)

  ws_liker = wb['좋아요']
  ws_comment = wb['댓글']

  sub_liker = ['번호','사용자ID','사용자명']
  for kwd, j in zip(sub_liker, list(range(1, len(sub_liker) + 1))):
    ws_liker.cell(row=1, column=j).value = kwd

  sub_comment = ['번호','레벨','사용자ID','내용']
  for kwd, j in zip(sub_comment, list(range(1, len(sub_comment) + 1))):
    ws_comment.cell(row=1, column=j).value = kwd

  wb.save(FILE_PATH + FILE_NAME)

# 엑셀 입력
def write_excel(crawl_type, append_row):
  wb = load_workbook(FILE_PATH + FILE_NAME, data_only=True)
  ws = wb.active
  if crawl_type == 'likers':
    ws = wb['좋아요']
  elif crawl_type == 'comments':
    ws = wb['댓글']

  ws.append(append_row)
  wb.save(FILE_PATH + FILE_NAME)

# 로그인 처리
def login_instagram(browser, id, password):
    browser.get(LOGIN_URL)
    sleep(5)
    inputs = browser.find_elements(by=By.TAG_NAME, value='input')
    inputs[0].clear()
    inputs[1].clear()
    inputs[0].send_keys(id)
    inputs[1].send_keys(password)
    inputs[1].submit()

# 게시글 주소에서 short code 추출
def get_short_code(target_url):
  short_code = ''
  try:
    short_code = target_url.split("/")[4]
  except Exception as e:
    short_code = ''
  finally:
    return short_code

def get_likers(browser, short_code):

  end_cursor = ''
  variables = {}
  variables['shortcode'] = short_code
  variables['first'] = 50
  variables['after'] = end_cursor
  has_next_page = True
  rownum = 0

  while (has_next_page):

    variables['after'] = end_cursor
    json_variables = str(json.dumps(variables))

    url = f'{LIKER_URL}&variables={json_variables}'

    browser.get(url)
    content = browser.find_element(by=By.TAG_NAME, value='pre').text
    data = json.loads(content)

    has_next_page = data['data']['shortcode_media']['edge_liked_by']['page_info']['has_next_page']
    end_cursor = data['data']['shortcode_media']['edge_liked_by']['page_info']['end_cursor']
    likers = data['data']['shortcode_media']['edge_liked_by']['edges']

    for liker in likers:
      rownum += 1
      write_excel('likers', [rownum, liker['node']['username'], unicodedata.normalize('NFC',liker['node']['full_name'])])
      # print(rownum, liker['node']['username'], liker['node']['full_name'])

def get_comments(browser, short_code):

  end_cursor = ''
  variables = {}
  variables['shortcode'] = short_code
  variables['first'] = 50
  variables['after'] = end_cursor
  has_next_page = True
  rownum = 0

  while (has_next_page):

    variables['after'] = end_cursor
    json_variables = str(json.dumps(variables))

    url = f'{COMMENT_URL}&variables={json_variables}'

    browser.get(url)
    content = browser.find_element(by=By.TAG_NAME, value='pre').text
    data = json.loads(content)

    has_next_page = data['data']['shortcode_media']['edge_media_to_parent_comment']['page_info']['has_next_page']
    end_cursor = data['data']['shortcode_media']['edge_media_to_parent_comment']['page_info']['end_cursor']
    comments = data['data']['shortcode_media']['edge_media_to_parent_comment']['edges']

    for comment in comments:
      rownum += 1
      write_excel('comments', [rownum, 1, comment['node']['owner']['username'], comment['node']['text']])
      # print(rownum, 1, comment['node']['owner']['username'], comment['node']['text'])
      co_comments = comment['node']['edge_threaded_comments']['edges']
      for co_comment in co_comments:
        write_excel('comments', [rownum, 2, co_comment['node']['owner']['username'], co_comment['node']['text']])
        # print(rownum, 2, co_comment['node']['owner']['username'], co_comment['node']['text'])

if __name__ == "__main__":
  app = QApplication(sys.argv)
  INSTA = INSTA_Window()
  INSTA.show()
  app.exec()
